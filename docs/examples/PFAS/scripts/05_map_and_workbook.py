#!/usr/bin/env python3
"""05_map_and_workbook.py -- interactive folium map fragment + multi-sheet xlsx."""
import json
import os
import sys

import numpy as np
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, "/sessions/busy-happy-edison/mnt/.claude/skills/okn-report-style/scripts")
from okn_figstyle import folium_osm_map, folium_map_iframe

D = os.path.abspath(os.path.join(HERE, "..", "data"))
ROOT = os.path.abspath(os.path.join(HERE, ".."))

S = json.load(open(os.path.join(D, "stats.json")))
ranked = pd.read_csv(os.path.join(D, "cells_ranked.tsv"), sep="\t", dtype={"cid": str})
top = pd.read_csv(os.path.join(D, "top20_named.tsv"), sep="\t", dtype={"cid": str})
named = pd.read_csv(os.path.join(D, "pfas_facilities_named.csv"), dtype=str)
ringn = pd.read_csv(os.path.join(D, "ring_pfas_facilities_named.csv"), dtype=str)

# ---------------------------------------------------------------- interactive map
TOPN = 150
m = ranked.head(TOPN).copy()
facs = []
for _, r in m.iterrows():
    s = list(named[named["cid"] == r["cid"]]["facName"].dropna().unique())[:3]
    g = list(ringn[ringn["cid"] == r["cid"]]["facName"].dropna().unique())[:3]
    facs.append("; ".join(s) if s else ("(adjacent cell) " + "; ".join(g) if g else "—"))
m["PFAS-flagged facilities"] = facs

rows = []
for _, r in m.iterrows():
    rows.append({
        "lat": r["lat"], "lon": r["lon"],
        "rank": int(r["rank"]),
        "score": round(float(r["score"]), 1),
        "tier": r["tier"],
        "county": f"{r['countyName']}, {r['stateName']}",
        "S2 cell (L13)": r["cid"],
        "detections / observations": f"{int(r['nDet'])} / {int(r['nObs'])}",
        "distinct analytes detected": int(r["nDetAnalytes"]),
        "max concentration (ng/L)": "n/a" if pd.isna(r["maxNgL"]) else f"{r['maxNgL']:,.1f}",
        "PFAS-flagged facilities (same cell)": int(r["nPfasFac"]),
        "PFAS-flagged facilities (1-ring)": int(r["nRingPfasFac"]),
        "all FRS facilities (same cell)": int(r["nFac"]),
        "nearest PFAS-flagged facilities": r["PFAS-flagged facilities"],
        "source KGs": r["sourceList"].replace(";", ", "),
    })

fmap = folium_osm_map(rows, value_key="score", tooltip_key="county", zoom_start=5,
                      popup_keys=[k for k in rows[0] if k not in ("lat", "lon")])
iframe = folium_map_iframe(fmap, height=540, title="Top-ranked PFAS sample cells")
with open(os.path.join(D, "map_fragment.html"), "w") as fh:
    fh.write(iframe)
print("map fragment written:", len(iframe), "chars")

# ---------------------------------------------------------------- workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TIER_FILL = {"A": "FDE0DC", "B": "FDF0D5", "C": "DEEBF7", "D": "F0F0F0", "N": "F7F7F7"}
HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10)

wb = Workbook()


def sheet(name, df, tier_col=None, widths=None, note=None):
    ws = wb.create_sheet(name[:31])
    start = 1
    if note:
        ws.cell(1, 1, note).font = Font(name="Arial", size=9, italic=True)
        ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(4, min(12, len(df.columns))))
        ws.row_dimensions[1].height = 30
        start = 3
    for j, c in enumerate(df.columns, 1):
        cell = ws.cell(start, j, str(c))
        cell.fill, cell.font = HDR, HDRF
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for i, (_, r) in enumerate(df.iterrows(), start + 1):
        fill = PatternFill("solid", fgColor=TIER_FILL[r[tier_col]]) if tier_col and r[tier_col] in TIER_FILL else None
        for j, c in enumerate(df.columns, 1):
            v = r[c]
            if isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating,)):
                v = None if pd.isna(v) else round(float(v), 4)
            cell = ws.cell(i, j, v)
            cell.font = BODY
            if fill:
                cell.fill = fill
    ws.freeze_panes = ws.cell(start + 1, 1)
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(df.columns))}{start + len(df)}"
    for j, c in enumerate(df.columns, 1):
        w = (widths or {}).get(c, min(38, max(10, int(df[c].astype(str).str.len().quantile(0.92)) + 3)))
        ws.column_dimensions[get_column_letter(j)].width = w
    return ws


RCOLS = ["rank", "cid", "score", "tier", "window", "stateName", "countyName", "countyFips",
         "nPts", "nObs", "nDet", "detFreq", "nDetAnalytes", "maxNgL",
         "nFac", "nPfasFac", "nRingFac", "nRingPfasFac",
         "sameTopIndustry", "sameTopNaics", "sameTopFacilityName",
         "ringTopIndustry", "ringTopNaics", "ringTopFacilityName",
         "c_proximity", "c_detIntensity", "c_detFreq", "c_analyteBreadth", "c_industryPrior",
         "nSources", "sourceList", "lat", "lon"]
sheet("Ranked Results", ranked[RCOLS], tier_col="tier",
      note="One row per S2 Level-13 cell (~1.3 km²) that contains a SAWGraph PFAS sample point AND at "
           "least one PFAS detection, ranked by co-location score. Row fill = confidence tier "
           "(A red, B amber, C blue, D grey). maxNgL is the maximum single-analyte aqueous "
           "concentration in ng/L; blank where the cell's detections were in non-aqueous media.")

neg = pd.read_csv(os.path.join(D, "cells_screened_negative.tsv"), sep="\t", dtype={"cid": str})
sheet("Screened Negative", neg[[c for c in RCOLS if c in neg.columns and c != "rank"]],
      note="Control set: cells screened for PFAS with observations but ZERO detections. "
           "Cells here that still carry a PFAS-flagged facility are the false positives of the "
           "co-location hypothesis.")

sheet("Analytes", pd.read_csv(os.path.join(D, "strat_analytes.tsv"), sep="\t"),
      note="Per-analyte detection statistics with the chemical crosswalk: CAS, EPA DSSTox identity, "
           "biobricks-ICE match, ToxCast assay-endpoint count, and ICE PREDICTED functional-use "
           "categories (QSUR model output — not curated OECD assignments).")
sheet("Industry Co-location", pd.read_csv(os.path.join(D, "strat_industry.tsv"), sep="\t"),
      note="PFAS-flagged industry groups co-located with PFAS sample cells, split by window "
           "(same cell vs 1-ring adjacent) with the source-strength prior applied in the score.")
sheet("NAICS detail", pd.read_csv(os.path.join(D, "strat_naics_same_cell.tsv"), sep="\t"),
      note="Leaf NAICS codes of EPA-PFAS-Facility-flagged facilities sharing a cell with a PFAS sample point.")
sheet("By State", pd.read_csv(os.path.join(D, "strat_state.tsv"), sep="\t"))
sheet("By County", pd.read_csv(os.path.join(D, "strat_county.tsv"), sep="\t"),
      note="Counties containing tier-A or tier-B cells.")
sheet("Functional Use", pd.read_csv(os.path.join(D, "strat_functional_use.tsv"), sep="\t"),
      note="Detection statistics grouped by ICE predicted functional-use category. "
           "Predicted (QSUR) assignments only — no PFAS in this set carries a curated OECD use.")
sheet("Statistical Tests", pd.read_csv(os.path.join(D, "tests.tsv"), sep="\t"))
sheet("Tier Summary", pd.read_csv(os.path.join(D, "strat_tier.tsv"), sep="\t"))

# ---- Methods & Rules ---------------------------------------------------------
meth = pd.DataFrame([
    ("Unit of analysis", "S2 Level-13 grid cell (~1.27 km² at the equator), the shared spatial key "
                         "linking sawgraph, fiokg and spatialkg."),
    ("Universe", f"{S['universe_cells']:,} S2 L13 cells containing ≥1 SAWGraph coso:SamplePoint."),
    ("Evaluable", f"{S['evaluable_cells']:,} of those also carry ≥1 analyte-linked contaminant observation."),
    ("Ranked set", f"{S['cells_with_detection']:,} cells with ≥1 detection (coso:DetectQuantityValue)."),
    ("Same-cell window", "A facility whose fiokg kwg:sfWithin cell equals the sample cell."),
    ("1-ring window", "A facility in any of the 8 S2 L13 cells adjacent to the sample cell, taken from "
                      "spatialkg spatial:connectedTo (verified: every cell has exactly 8 neighbours)."),
    ("Tier A", "Detection + ≥1 EPA-PFAS-Facility in the SAME cell."),
    ("Tier B", "Detection + ≥1 EPA-PFAS-Facility in the 1-ring, none in the same cell."),
    ("Tier C", "Detection + ≥1 FRS facility in the window, but none PFAS-flagged."),
    ("Tier D", "Detection with no regulated facility in the window."),
    ("Tier N", "Screened with observations but zero detections (control set, not ranked)."),
    ("Tier X", f"{S['tierX']:,} cells with a sample point but no analyte-linked observation — excluded."),
    ("Score", "100 × Σ wᵢcᵢ / Σ wᵢ over AVAILABLE components. Weights: proximity 0.35, detection "
              "intensity 0.25, detection frequency 0.20, analyte breadth 0.10, industry prior 0.10."),
    ("c_proximity", "max(P_same, 0.60·P_ring, 0.25·F_same, 0.10·F_ring) where P/F are log1p-saturating "
                    "transforms of the PFAS-flagged / all-facility counts (saturation 5, 10, 20, 40)."),
    ("c_detIntensity", "min(1, log10(1+maxNgL)/log10(1001)); undefined where no ng/L detection exists."),
    ("c_detFreq", "detections ÷ observations in the cell."),
    ("c_analyteBreadth", "min(1, distinct analytes detected ÷ 20)."),
    ("c_industryPrior", "Source-strength prior of the strongest co-located PFAS-flagged industry: "
                        "High 1.0, Moderate 0.6, Low 0.3, none 0."),
    ("Chemical crosswalk", "sawgraph coso:casNumber (undashed digits → dashed) → "
                           "http://identifiers.org/cas/{cas} → edam:has_identifier in biobricks-ice / "
                           "biobricks-toxcast."),
    ("Abbreviations", "PFAS perfluoroalkyl and polyfluoroalkyl substances · AFFF aqueous film-forming foam · "
                      "FRS EPA Facility Registry Service · NAICS North American Industry Classification "
                      "System · EGAD Maine DEP Environmental & Geographic Analysis Database · WQP Water "
                      "Quality Portal · ICE Integrated Chemical Environment · QSUR quantitative "
                      "structure–use relationship · DTXSID DSSTox substance identifier · CAS Chemical "
                      "Abstracts Service registry number · WWTF wastewater treatment facility · "
                      "S2 Google S2 spherical geometry grid · KG knowledge graph."),
], columns=["Rule", "Definition"])
sheet("Methods & Rules", meth, widths={"Rule": 26, "Definition": 110})

del wb["Sheet"]
out = os.path.join(ROOT, "pfas_source_attribution_results.xlsx")
wb.save(out)
print("workbook:", out, os.path.getsize(out), "bytes;", len(wb.sheetnames), "sheets")
