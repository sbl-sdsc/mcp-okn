#!/usr/bin/env python3
"""Build study/MS_results.xlsx — the 12-sheet results workbook for the OKN MS study.

Arial throughout, filled + frozen + auto-filtered header row, wrapped text, sized columns.
Run:  python3 make_workbook.py
"""
from __future__ import annotations

import collections
import json
import re
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, "/sessions/vibrant-kind-heisenberg/mnt/outputs/ms_work/study/scripts")
from make_figures import band_label, contrast_meta, normalise_drug  # noqa: E402

WD = "/sessions/vibrant-kind-heisenberg/mnt/outputs/ms_work"
DATA = f"{WD}/data"
OUT = f"{WD}/study/MS_results.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT, size=10)
TIER_FILL = {"A": PatternFill("solid", fgColor="C6EFCE"),
             "B": PatternFill("solid", fgColor="FFE699"),
             "C": PatternFill("solid", fgColor="F2F2F2")}
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=THIN)


def write_sheet(wb, title, df, widths=None, wrap_cols=(), tier_col=None, freeze="A2",
                number_formats=None):
    ws = wb.create_sheet(title)
    cols = list(df.columns)
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for _, row in df.iterrows():
        ws.append(["" if pd.isna(v) else v for v in row.tolist()])
    nrow, ncol = len(df) + 1, len(cols)
    for r in range(2, nrow + 1):
        tier = df.iloc[r - 2][tier_col] if tier_col else None
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            name = cols[c - 1]
            cell.alignment = Alignment(vertical="top", wrap_text=name in wrap_cols)
            if tier and tier in TIER_FILL:
                cell.fill = TIER_FILL[tier]
            if number_formats and name in number_formats:
                cell.number_format = number_formats[name]
    if nrow > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{nrow}"
    ws.freeze_panes = freeze
    ws.row_dimensions[1].height = 30
    for i, name in enumerate(cols, start=1):
        if widths and name in widths:
            w = widths[name]
        else:
            sample = df[name].astype(str).head(400)
            w = min(max(12, int(sample.map(len).max() if len(sample) else 12) + 2,
                        len(str(name)) + 2), 46)
        ws.column_dimensions[get_column_letter(i)].width = w
    print(f"  {title:28s} {len(df):5d} rows x {ncol} cols")
    return ws


def sci(x):
    return "0.00E+00"


# ---------------------------------------------------------------- 1. Ranked Results
def sheet_ranked():
    g = pd.read_csv(f"{DATA}/gene_evidence_master.csv")
    g["tier_rank"] = g.tier.map({"A": 0, "B": 1, "C": 2})
    g = g.sort_values(["tier_rank", "consensus_score", "n_evidence_types", "n_sources", "gene"],
                      ascending=[True, False, False, False, True]).drop(columns=["tier_rank"])
    g = g[["gene", "tier", "consensus_score", "n_evidence_types", "n_sources", "evidence_types",
           "sources", "pigean_score", "de_assays", "de_direction", "de_contexts"]]
    g.columns = ["gene symbol", "tier", "consensus score", "n evidence types", "n source KGs",
                 "evidence types", "source KGs", "PIGEAN score", "n DE assays", "DE direction",
                 "DE contexts"]
    return g


# ---------------------------------------------------------------- 2. Identifier Crosswalk
CROSSWALK = [
    ("MONDO_0005301", "multiple sclerosis", "DOID:2377", "", "UMLS:C0026769", "MESH:D009103",
     "NCIT:C3243", "Orphanet:802", "ICD10CM:G35", "ICD9:340", "SCTID:24700007", "",
     "MEDGEN:10123",
     "spoke-okn (as DOID_2377); prokn (DOID_2377 + MeSH D009103 node); rdkg; "
     "gene-expression-atlas-okn; biomarkerkg; ubergraph"),
    ("MONDO_0005314", "relapsing-remitting multiple sclerosis", "DOID:2378", "EFO:0003929",
     "UMLS:C0751967", "", "NCIT:C165675", "", "", "", "SCTID:426373005", "", "MEDGEN:155669",
     "prokn (DOID_2378); rdkg; gene-expression-atlas-okn (EFO_0003929); ubergraph"),
    ("MONDO_0000450", "secondary progressive multiple sclerosis", "DOID:0050783", "EFO:0008522",
     "UMLS:C0751965", "", "", "", "", "", "SCTID:425500002", "", "MEDGEN:155969",
     "prokn (DOID_0050783); gene-expression-atlas-okn (EFO_0008522); ubergraph"),
    ("MONDO_0000451", "primary progressive multiple sclerosis", "DOID:0050784", "EFO:0008520",
     "UMLS:C0751964", "", "", "", "", "", "SCTID:428700003", "", "MEDGEN:155968",
     "gene-expression-atlas-okn (EFO_0008520); ubergraph"),
    ("MONDO_0000452", "progressive relapsing multiple sclerosis", "DOID:0050785", "",
     "UMLS:C0393666", "", "", "", "", "", "SCTID:230374002", "", "MEDGEN:95982",
     "rdkg; ubergraph"),
    ("MONDO_0005284", "chronic progressive multiple sclerosis", "", "EFO:0003840",
     "UMLS:C0393665", "MESH:D020528", "", "", "", "", "SCTID:230373008", "", "MEDGEN:140733",
     "ubergraph"),
    ("MONDO_0016429", "Marburg acute multiple sclerosis", "", "", "UMLS:C4707723", "", "",
     "Orphanet:228157", "", "", "SCTID:766246000", "GARD:20572", "MEDGEN:1641985",
     "oard-kg; ubergraph"),
    ("MONDO_0018784", "pediatric multiple sclerosis", "", "", "UMLS:C5568571", "", "",
     "Orphanet:477738", "", "", "", "GARD:10443", "MEDGEN:1799994",
     "rdkg; oard-kg; ubergraph"),
    ("— (no MONDO term)", "clinically isolated syndrome (CIS)", "", "", "", "", "", "", "", "",
     "", "", "",
     "NOTE: Clinically Isolated Syndrome has NO MONDO term under multiple sclerosis "
     "(MONDO_0005301), so it cannot appear on the subtype axis of this study and is absent "
     "from every subtype-resolved result (figures 8A and the Therapeutics sheet)."),
]


def sheet_crosswalk():
    cols = ["MONDO", "label", "DOID", "EFO", "UMLS", "MeSH", "NCIT", "Orphanet", "ICD-10",
            "ICD-9", "SNOMED CT", "GARD", "MedGen", "present in which KGs"]
    return pd.DataFrame(CROSSWALK, columns=cols)


# ---------------------------------------------------------------- 6/7. GXA sheets
def sheet_de():
    g, meta = contrast_meta()
    rows = []
    for r in g["genes"]:
        m = meta[r["assay"]]
        rows.append({
            "gene symbol": r["sym"], "NCBI Gene": r["gene"],
            "direction in MS": r["dir"], "contrast (assay name)": m["name"],
            "contrast (short)": m["short"], "compartment / cell type": m["comp"],
            "MS subtype": m["sub"], "study": m["study"], "technology": m["tech"],
            "assay IRI": r["assay"], "disease term": r["d"].rsplit("/", 1)[-1],
        })
    return pd.DataFrame(rows).sort_values(
        ["compartment / cell type", "contrast (short)", "direction in MS", "gene symbol"])


def sheet_gxa_enrich():
    g, meta = contrast_meta()
    a2 = {m["name"]: m for m in meta.values()}
    rows = []
    for r in g["enrichment"]:
        m = meta.get(r["assay"]) or a2.get(r["assayName"], {})
        rows.append({
            "source": r["src"], "term id": r["term"].rsplit("/", 1)[-1].split("#")[-1],
            "term name": r["termName"], "direction": r["dir"],
            "effect size (KG)": r["es"], "p-value (KG)": r["pv"],
            "contrast (assay name)": r["assayName"],
            "contrast (short)": m.get("short", ""),
            "compartment / cell type": m.get("comp", ""), "MS subtype": m.get("sub", ""),
            "study": r["study"],
        })
    return pd.DataFrame(rows).sort_values(["source", "term name", "contrast (short)"])


# ---------------------------------------------------------------- 9. Therapeutics
def sheet_therapeutics():
    r = pd.read_csv(f"{DATA}/rdkg_ms_drugs.csv")
    r["norm"] = [normalise_drug(x) for x in r.drugLabel]
    sub_lab = {"MONDO_0005301": "MS", "MONDO_0005314": "RRMS",
               "MONDO_0000452": "PRMS", "MONDO_0018784": "pediatric MS"}
    rd = collections.defaultdict(lambda: {"sub": set(), "rel": set(), "raw": set(), "id": set()})
    for _, x in r.iterrows():
        k = x.norm.lower()
        rd[k]["sub"].add(sub_lab[x.mondo])
        rd[k]["rel"].add(x.rel)
        rd[k]["raw"].add(x.drugLabel)
        rd[k]["id"].add(x.drugId)
        rd[k]["name"] = x.norm
    p = pd.read_csv(f"{DATA}/ms_drugs_targets.csv")
    pk = {}
    for _, x in p.iterrows():
        pk[str(x.drug).lower()] = x
    order = ["MS", "RRMS", "PRMS", "pediatric MS"]
    rows = []
    for k in sorted(set(rd) | set(pk)):
        d = rd.get(k)
        q = pk.get(k)
        rows.append({
            "drug": (d["name"] if d else str(q.drug)),
            "in rdkg": "yes" if d else "no",
            "rdkg MS subtypes treated": ", ".join(s for s in order if d and s in d["sub"]) if d else "",
            "rdkg relation": ", ".join(sorted(d["rel"])) if d else "",
            "rdkg source labels (trial arms merged)": "; ".join(sorted(d["raw"])) if d else "",
            "rdkg drug id": "; ".join(sorted(d["id"])) if d else "",
            "in prokn": "yes" if q is not None else "no",
            "prokn MS indication terms": (q.ms_indication_terms if q is not None else ""),
            "prokn n targets": (int(q.n_targets) if q is not None and not pd.isna(q.n_targets) else ""),
            "prokn targets": (q.targets if q is not None else ""),
            "prokn evidence relation": (q.evidence_relations if q is not None else ""),
            "prokn max phase": (q.phase if q is not None else ""),
            "prokn approval year": (q.approval_year if q is not None else ""),
        })
    df = pd.DataFrame(rows)
    df["_o"] = (df["in rdkg"] == "no").astype(int)
    return df.sort_values(["_o", "drug"]).drop(columns="_o")


# ---------------------------------------------------------------- 11. Source Inventory
INVENTORY = [
    ("ubergraph", "v0.0.2", "2026-05-01", "MONDO/DOID/EFO subtype closure and identifier "
     "cross-references for the MS subtree", "MONDO / DOID / EFO IRI",
     "USED — supplied the 8-term MS subtree and 105 cross-references"),
    ("spoke-okn", "v0.0.6", "2026-03-16", "curated disease-gene edges for MS; country-level "
     "prevalence and mortality (IHME GBD 2019)", "DOID_2377; ISO3 country code",
     "USED — 164 disease genes, 200 prevalence rows, 178 mortality rows"),
    ("rdkg", "v0.0.1", "2026-05-04", "subtype-resolved drug→disease treats / "
     "contraindicated_for assertions (DrugBank + clinical-trial arms)", "MONDO IRI; DrugBank id",
     "USED — 93 rows over 4 MS subtypes, 78 distinct drug nodes"),
    ("prokn", "v0.0.5", "2026-06-23", "drug indications for MS, drug→protein target edges, and "
     "the GO/Reactome annotation background for enrichment", "gene symbol; UniProt; DOID/MeSH",
     "USED — 249 MS-indicated drugs, 94 targets, 2,321 annotation rows, enrichment background"),
    ("digcfdekg", "v0.0.1", "2026-06-21", "PIGEAN gene-trait scores for MS and MS MRI endpoints",
     "gene symbol", "USED — 1,548 gene-trait rows over 1,059 genes (+3 MRI traits)"),
    ("gene-expression-atlas-okn", "v0.0.3", "2026-03-18", "differential expression and "
     "contrast-level GO/Reactome/InterPro enrichment for MS contrasts",
     "gene symbol / NCBI Gene; MONDO / EFO", "USED — 19 MS contrasts, 790 DE rows, "
     "332 enrichment rows"),
    ("biomarkerkg", "v0.0.2", "2026-03-16", "clinical biomarker assertions for MS "
     "(risk / diagnostic / prognostic)", "gene symbol; MONDO_0005301",
     "USED — 394 rows, 373 distinct biomarkers, 257 genes"),
    ("oard-kg", "v0.0.3", "2026-06-05", "phenotype (HPO) associations for rare MS subtypes",
     "MONDO IRI; HP id", "USED — 211 phenotype rows (209 pediatric MS, 2 Marburg); "
     "no rows for adult MS"),
    ("biohealth", "v0.0.4", "2026-03-16", "literature-derived comorbidity and social-determinant "
     "context for MS concepts", "UMLS CUI",
     "USED — 12,033 edges over 5 MS CUIs, 460 SDoH provenance statements"),
    ("nde", "v0.0.3", "2026-03-16", "NIAID Data Ecosystem dataset metadata mentioning MS",
     "MONDO / DOID health-condition term",
     "USED — 853 dataset records (2 with an Epstein-Barr virus organism link)"),
    ("ncipidkg", "v0.0.1", "2026-04-03", "NCI-PID protein-interaction context for MS core "
     "proteins", "UniProt accession",
     "DROPPED — clean negative: the published graph is a small demonstration subgraph "
     "(120 statements) with no overlap with the MS core proteins"),
    ("pankgraph", "v0.0.1", "2026-03-23", "islet gene/variant context; probed for MS "
     "gene-disease edges", "Ensembl gene id; MONDO IRI",
     "DROPPED for MS — its disease axis contains exactly one disease (type 1 diabetes); "
     "0 MS terms. (All 59 MS core genes exist as gene nodes, so it is retained only as an "
     "islet/T1D comparator, contributing no MS evidence row.)"),
    ("evoweb", "v0.0.2", "2026-06-04", "protein co-evolution clusters; probed for MS proteins",
     "protein accession",
     "DROPPED — definitive negative: 100% of its 3,269,864 members are prokaryotic RefSeq WP_ "
     "accessions; zero human proteins, no join key"),
    ("wikidata", "no VoID provenance", "no VoID provenance",
     "country centroid coordinates (P625) for the epidemiology map", "ISO3 country code",
     "USED — 200 country centroids; wikidata records no VoID version"),
]


def sheet_inventory():
    return pd.DataFrame(INVENTORY, columns=[
        "knowledge graph", "version", "last updated", "role in this study",
        "join key", "used / dropped + reason"])


# ---------------------------------------------------------------- 12. Methods & Rules
METHODS = [
    ("Scope", "Disease anchor",
     "MONDO_0005301 'multiple sclerosis' plus its complete ubergraph rdfs:subClassOf* closure "
     "(8 terms). Clinically Isolated Syndrome (CIS) has no MONDO term under multiple sclerosis "
     "and therefore never appears on the subtype axis."),
    ("Scope", "Knowledge graphs queried",
     "14 (ubergraph, spoke-okn, rdkg, prokn, digcfdekg, gene-expression-atlas-okn, biomarkerkg, "
     "oard-kg, biohealth, nde, ncipidkg, pankgraph, evoweb, wikidata); 11 contributed evidence "
     "rows, 3 were queried and dropped (see Source Inventory)."),
    ("Enrichment", "Test",
     "One-sided hypergeometric over-representation test (sf(k-1, N, K, n)), Benjamini–Hochberg "
     "FDR correction across the terms tested within each annotation family; significance at "
     "FDR < 0.05."),
    ("Enrichment", "Foreground",
     "The MS evidence gene set intersected with the genes carrying that annotation type in "
     "prokn: n = 86 (GO BP), n = 65 (GO MF), n = 66 (GO CC), n = 71 (Reactome)."),
    ("Enrichment", "Background — GO biological process",
     "N = 7,663 distinct prokn gene symbols carrying at least one GO BP annotation."),
    ("Enrichment", "Background — GO molecular function",
     "N = 8,033 distinct prokn gene symbols carrying at least one GO MF annotation."),
    ("Enrichment", "Background — GO cellular component",
     "N = 8,094 distinct prokn gene symbols carrying at least one GO CC annotation."),
    ("Enrichment", "Background — Reactome",
     "N = 6,032 distinct prokn gene symbols carrying at least one Reactome pathway annotation."),
    ("Enrichment", "Result counts",
     "GO BP 118 / 135 terms significant; GO MF 24 / 43; GO CC 12 / 32; Reactome 43 / 47."),
    ("Enrichment", "Column key",
     "k = MS genes in the term; K = background genes in the term; n = MS genes tested; "
     "N = background size; expected = n·K/N; fold = k / expected."),
    ("Tiering", "Consensus tier rule",
     "Tier A = the gene is supported by >= 4 independent evidence types; Tier B = exactly 3; "
     "Tier C = <= 2. Result: A n = 52, B n = 80, C n = 2,265 (2,397 genes total)."),
    ("Tiering", "Evidence types counted",
     "curated_disease_gene (spoke-okn), genetic_association (digcfdekg PIGEAN), "
     "genetic_association_mri_endpoint (digcfdekg MRI traits), differential_expression "
     "(gene-expression-atlas-okn), clinical_biomarker (biomarkerkg), pathway_go_membership "
     "(prokn). A type is counted once however many rows support it."),
    ("Tiering", "Consensus score",
     "consensus_score = 2·(n evidence types) + (n source KGs); used only to order genes within "
     "a tier, never to move a gene between tiers."),
    ("Differential expression", "Threshold",
     "The up / down calls published by gene-expression-atlas-okn on the gene→assay edge were "
     "used as-is. The KG exposes no effect size or p-value on that gene-level edge, so this "
     "study applied no further numeric threshold. 5 of the 19 MS contrasts carry gene-level DE "
     "rows (790 rows, 777 distinct genes); 13 of 19 carry contrast-level enrichment."),
    ("Differential expression", "Contrast-level enrichment",
     "GXA contrast enrichment rows (GO / Reactome / InterPro) carry the KG's own effect size "
     "and p-value and were retained as published (all 332 rows have p <= 1.9e-04); they are NOT "
     "re-tested or FDR-corrected by this study."),
    ("Epidemiology", "Source and statistics",
     "spoke-okn PREVALENCE_DpL (IHME GBD 2019) for 200 countries, joined to wikidata P625 "
     "country centroids on ISO3. Spearman rank correlation of |latitude| against prevalence "
     "per 100,000: rho = 0.836, p = 2.1e-53, n = 200 (Northern hemisphere rho = 0.845, "
     "p = 3.4e-44, n = 158; Southern rho = 0.679, p = 7.5e-07, n = 42)."),
    ("Epidemiology", "Caveat",
     "Country centroid coordinates are a coarse proxy for population latitude; prevalence is a "
     "modelled GBD estimate with a 95% uncertainty interval (columns lo / hi), and the "
     "association is ecological and observational, not causal."),
    ("Therapeutics", "rdkg drug normalisation",
     "rdkg trial-arm labels were merged onto a single agent for the subtype matrix "
     "(e.g. 'Ocrelizumab Dose 1', 'Short-course Ocrelizumab', 'Continued Ocrelizumab' -> "
     "'Ocrelizumab'; 'Copaxone®' -> 'Glatiramer acetate'; 'Tysabri' -> 'Natalizumab'; "
     "'Bg00012' -> 'Dimethyl fumarate'): 78 raw labels -> 55 agents."),
    ("Therapeutics", "Target ranking",
     "Targets were exploded from the prokn ms_drugs_targets 'targets' column; bare CHEMBL* "
     "identifiers were excluded. Ties in the top-20 ranking are broken in favour of the eight "
     "established MS drug targets (S1PR1, S1PR5, MS4A1, DHODH, KEAP1, BTK, TOP2A, NR3C1), "
     "then alphabetically."),
    ("Standing caveats", "Association is not causation",
     "Every gene-disease, biomarker, comorbidity and SDoH edge in this study is an "
     "observational association asserted by a source KG (much of it literature-mined), not "
     "evidence of causation."),
    ("Standing caveats", "Provenance",
     "Counts are reproducible against the logged SPARQL queries and the KG versions recorded "
     "in the Source Inventory sheet; re-running against a later KG release may change them."),
    ("Abbreviations", "Abbreviations",
     "MS multiple sclerosis · CIS clinically isolated syndrome · RRMS relapsing-remitting MS · "
     "SPMS secondary progressive MS · PPMS primary progressive MS · PRMS progressive relapsing "
     "MS · DMT disease-modifying therapy · GWAS genome-wide association study · FDR "
     "false-discovery rate · GO Gene Ontology · BP biological process · MF molecular function · "
     "CC cellular component · DE differential expression · GXA Expression Atlas "
     "(gene-expression-atlas-okn) · KG knowledge graph · SDoH social determinants of health · "
     "EBV Epstein-Barr virus · NfL neurofilament light chain · OCB oligoclonal bands · "
     "IHME Institute for Health Metrics and Evaluation · GBD Global Burden of Disease · "
     "HP / HPO Human Phenotype Ontology · MONDO Mondo Disease Ontology · DOID Human Disease "
     "Ontology · EFO Experimental Factor Ontology · UMLS Unified Medical Language System"),
]


def sheet_methods():
    return pd.DataFrame(METHODS, columns=["section", "item", "value"])


# ---------------------------------------------------------------- main
def main():
    wb = Workbook()
    wb.remove(wb.active)
    print("building", OUT)

    write_sheet(wb, "Ranked Results", sheet_ranked(),
                widths={"evidence types": 44, "source KGs": 40, "DE contexts": 46,
                        "DE direction": 16, "gene symbol": 14},
                wrap_cols={"evidence types", "source KGs", "DE contexts"}, tier_col="tier")

    write_sheet(wb, "Identifier Crosswalk", sheet_crosswalk(),
                widths={"label": 38, "present in which KGs": 60, "MONDO": 18},
                wrap_cols={"label", "present in which KGs"})

    bp = pd.read_csv(f"{DATA}/enrichment_go_bp.csv")
    bp.columns = ["GO id", "term", "k", "K", "n", "N", "expected", "fold", "p", "FDR"]
    write_sheet(wb, "GO BP Enrichment", bp, widths={"term": 58},
                wrap_cols={"term"}, number_formats={"p": sci(0), "FDR": sci(0)})

    mf = pd.read_csv(f"{DATA}/enrichment_go_mfcc.csv")
    mf.columns = ["aspect", "GO id", "term", "k", "K", "n", "N", "expected", "fold", "p", "FDR"]
    write_sheet(wb, "GO MF-CC Enrichment", mf, widths={"term": 58},
                wrap_cols={"term"}, number_formats={"p": sci(0), "FDR": sci(0)})

    rc = pd.read_csv(f"{DATA}/enrichment_reactome.csv")
    rc.columns = ["Reactome id", "pathway", "k", "K", "n", "N", "expected", "fold", "p", "FDR"]
    write_sheet(wb, "Reactome Enrichment", rc, widths={"pathway": 58},
                wrap_cols={"pathway"}, number_formats={"p": sci(0), "FDR": sci(0)})

    write_sheet(wb, "Differential Expression", sheet_de(),
                widths={"contrast (assay name)": 56, "contrast (short)": 40, "assay IRI": 46,
                        "NCBI Gene": 40},
                wrap_cols={"contrast (assay name)", "contrast (short)"})

    write_sheet(wb, "GXA Contrast Enrichment", sheet_gxa_enrich(),
                widths={"term name": 46, "contrast (assay name)": 56, "contrast (short)": 40},
                wrap_cols={"term name", "contrast (assay name)", "contrast (short)"},
                number_formats={"p-value (KG)": sci(0)})

    bm = pd.read_csv(f"{DATA}/biomarkers_table.csv")
    bm.columns = ["biomarker id", "biomarker label", "relation to MS", "gene",
                  "assessed entity", "sample type"]
    write_sheet(wb, "Biomarkers", bm, widths={"biomarker label": 52, "assessed entity": 40},
                wrap_cols={"biomarker label", "assessed entity"})

    write_sheet(wb, "Therapeutics", sheet_therapeutics(),
                widths={"drug": 34, "rdkg source labels (trial arms merged)": 52,
                        "prokn targets": 52, "rdkg MS subtypes treated": 24},
                wrap_cols={"rdkg source labels (trial arms merged)", "prokn targets", "drug"})

    ep = pd.read_csv(f"{DATA}/ms_prevalence_latitude.csv")
    ep["abslat band (°)"] = ep.abslat.map(band_label)
    ep = ep[["iso3", "country", "per100k", "lo", "hi", "prev_pct", "lat", "lon", "abslat",
             "abslat band (°)", "year"]]
    ep.columns = ["ISO3", "country", "prevalence per 100k", "95% UI lower", "95% UI upper",
                  "prevalence (%)", "latitude", "longitude", "|latitude|", "abslat band (°)",
                  "GBD year"]
    write_sheet(wb, "Epidemiology", ep.sort_values("prevalence per 100k", ascending=False),
                number_formats={"prevalence per 100k": "0.00", "95% UI lower": "0.00",
                                "95% UI upper": "0.00", "prevalence (%)": "0.0000"})

    write_sheet(wb, "Source Inventory", sheet_inventory(),
                widths={"role in this study": 52, "used / dropped + reason": 62,
                        "join key": 26, "knowledge graph": 26},
                wrap_cols={"role in this study", "used / dropped + reason", "join key"})

    write_sheet(wb, "Methods & Rules", sheet_methods(),
                widths={"section": 22, "item": 34, "value": 110},
                wrap_cols={"value", "item"})

    wb.save(OUT)
    print("saved", OUT)


if __name__ == "__main__":
    main()
