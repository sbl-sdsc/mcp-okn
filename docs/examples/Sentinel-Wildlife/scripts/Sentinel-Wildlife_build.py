#!/usr/bin/env python3
"""Sentinel-Wildlife: build the XLSX workbook and render the interactive HTML report."""
import json, sys, pathlib, ast
import pandas as pd
sys.path.insert(0, "/sessions/eager-brave-pascal/mnt/.claude/skills/okn-report-style/scripts")
from build_report_html import (candidate_table, build_report_from_markdown,
                               kpis_from_stats, fill_stats, load_stats)
D = "data"
st = load_stats("data/stats.json")
c  = pd.read_csv(f"{D}/county_priority_ranking.csv", dtype={"fips5": str})
sp = pd.read_csv(f"{D}/species_priority_ranking.csv")

# ------------------------------------------------------------------ XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
TIERFILL = {"A": "FFE2C8", "B": "FFF6DA", "C": "EDEFF0"}
wb = Workbook(); wb.remove(wb.active)

def sheet(name, df, tier_col=None, widths=None, wrap=()):
    ws = wb.create_sheet(name[:31])
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F4858")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for _, r in df.iterrows():
        ws.append(list(r.values))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            if cell.column_letter in wrap:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if tier_col:
            t = str(row[list(df.columns).index(tier_col)].value)
            if t in TIERFILL:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=TIERFILL[t])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(df.columns, start=1):
        w = (widths or {}).get(col, min(max(12, int(df[col].astype(str).str.len().max() or 12) + 2), 46))
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

cc = c[["rank", "county_name", "fips5", "conf_tier", "priority", "sentinel_species",
        "best_tier", "host_species", "total_species", "species_n_bird", "species_n_amphibian",
        "records_bird", "records_amphibian", "pfas_facilities", "frs_facilities", "spoke_chemicals_found",
        "adult_asthma_pct", "per_capita_income", "cm_cities", "climate_papers",
        "pfas_samples_in_county", "bridge", "lat", "lon", "sources_n", "evidence"]].copy()
cc.columns = ["rank", "county", "FIPS", "tier", "priority score", "sentinel-capable species",
              "best proximity tier", "pathogen-host species", "species observed",
              "bird species", "amphibian species", "bird records", "amphibian records",
              "EPA PFAS facilities", "EPA FRS facilities (all)", "spoke-okn chemicals in county", "adult asthma %",
              "per-capita income USD", "climatemodelskg places", "climate papers",
              "contaminant samples in county", "county bridge", "anchor lat", "anchor lon",
              "supporting KGs", "evidence"]
sheet("Ranked Results", cc, tier_col="tier", wrap=("Y",))

ss = sp[["rank", "binom", "taxon_level", "conf_tier", "value", "tier", "tier_label", "anchor",
         "measured", "pathogen_diseases", "biohealth_edges", "clinical_biomarker_items",
         "nde_datasets", "fl_records", "fl_individuals", "fl_counties", "gap",
         "artefact_flag", "sources_n", "evidence"]].copy()
ss.columns = ["rank", "species", "taxon level", "tier", "information value",
              "proximity tier", "proximity tier meaning", "clade anchor",
              "body burden measured", "infectious diseases hosted",
              "biohealth human-evidence edges", "EHR phenotypes + biomarkers",
              "nde datasets", "FL records", "FL individuals", "FL counties",
              "gap (host, never sampled)", "data-quality flag", "supporting KGs", "evidence"]
sheet("Species Ranking", ss, tier_col="tier", wrap=("T",))

sheet("Observation Inventory", pd.read_csv(f"{D}/wl_species_summary.csv"))
sheet("Observations by County", pd.read_csv(f"{D}/wl_county_clade.csv"))
sheet("Temporal Series", pd.read_csv(f"{D}/wl_obs_by_year.csv"))
sheet("Contaminant Record", pd.read_csv(f"{D}/sawgraph_biota_taxa.csv"))
sheet("Contaminant Geography", pd.read_csv(f"{D}/sawgraph_sampling_states.csv", dtype=str))
sheet("Avian PFAS Measurements", pd.read_csv(f"{D}/bird_pfas_detects.csv"))
sheet("Proximity Tiers", pd.read_csv(f"{D}/phylo_tiers.csv"))
sheet("Host-Pathogen Links", pd.read_csv(f"{D}/host_pathogen_biohealth.csv"))
sheet("NDE Taxon Overlap", pd.read_csv(f"{D}/nde_taxon_overlap.csv"))
sheet("Human Evidence", pd.read_csv(f"{D}/human_evidence_by_disease.csv"))
sheet("Place Context", pd.read_csv(f"{D}/fl_place_context.csv", dtype={"fips5": str}))
sheet("County Bridge", pd.concat([
    pd.read_csv(f"{D}/fl_county_bridge.csv", dtype=str).assign(note="verified L8 label bridge"),
    pd.read_csv(f"{D}/fl_county_bridge_repair.csv", dtype=str)]))

methods = pd.DataFrame([
 ("Study area", f"{st['fl_counties_studied']} Florida counties: {st['l8_true_count']} via the verified wildlifekn x spatialkg county-name bridge + 2 recovered by normalising 'Saint' to 'St.'. City-type wildlifekn locations (757 records) are excluded from all county statements."),
 ("Measured vs inferred", "MEASURED = a sawgraph contaminant concentration in a biota sample of that exact NCBITaxon id. INFERRED = shares a genus (I1), subfamily (I2), family (I3) or superorder (I4) with a measured species via ubergraph rdfs:subClassOf*. N = nearest measured relative only at class Aves. Z = no measured relative anywhere in Amphibia."),
 ("County score", "min-max normalised weighted sum: 0.30 sentinel-capable species + 0.20 best proximity tier weight + 0.20 pathogen-host species + 0.12 log10(EPA FRS facilities) + 0.13 total observed species + 0.05 adult asthma %. Tier A = top quartile, B = 40th-75th pct, C = below 40th pct."),
 ("Species score", "min-max normalised weighted sum: 0.28 infectious diseases hosted + 0.22 proximity tier weight + 0.18 biohealth human-evidence edges + 0.16 FL records + 0.10 FL counties + 0.06 has nde datasets; multiplied by 1.35 when the species is a pathogen host with no measured body burden. Higher taxa (Anura, Amphibia) are excluded from the ranking."),
 ("Environmental-pressure proxy", "fiokg EPA-PFAS-Facility records per county (4,118 in Florida; 188,057 nationally) — the PFAS-relevant subset of the EPA Facility Registry Service. All-FRS counts (118,663 in Florida of 4,955,792 nationally) are reported as a column but do NOT enter the score: FRS registers every site ever regulated or reported by EPA or a state programme, so it indexes economic activity more than PFAS risk. Verified 2026-07-29: log10 of the two measures correlates r = 0.943 across the 64 study counties, the top 8 counties are identical under either proxy, and the median absolute rank shift is 1 (max 12)."),
 ("Sampling deficit", "Uniform: every Florida county has zero contaminant samples of any medium, so the deficit term does not discriminate within the study area."),
 ("Infectious-disease filter", "nde MONDO health conditions restricted by ubergraph rdfs:subClassOf* under MONDO:0005550 (infectious disease)."),
 ("Human evidence", "biohealth via ubergraph MONDO oboInOwl:hasDbXref -> UMLS CUI node; oard-kg reified associations UNIONing biolink:subject and biolink:object; biomarkerkg OBCI_1000008/OBCI_1000002; prokn skos:exactMatch; spoke-okn DOID node IRIs via ubergraph skos:exactMatch."),
 ("Declared skips", "GO and Reactome enrichment and drug/target linkage were not run: the study has no gene or protein foreground. biobricks-aopwiki (taxonomic applicability of adverse outcome pathways), rdkg, digcfdekg, gene-expression-atlas-okn, spoke-genelab, biobricks-mesh and spoke-okn organisms were considered and dropped with reasons in report section 6.4."),
 ("Known defects surfaced", f"Crosswalk L8 published verified_count is {st['l8_published_count']}; the true Florida county count is {st['l8_true_count']} (one distinct value is the literal string 'https'). wildlifekn x sawgraph clade-expanded taxon overlap of 339 is an artefact of materialised phylum-level ancestors and must not be read as 339 species with contaminant data. biohealth asserts Anura->influenza (text-mining artefact) and misspells Colinus virginianus as 'Colinus virginiuanus' (silent false negative in the label bridge)."),
 ("Caveat carried to every claim", "Hypothesis generation for sampling design; not exposure assessment, not causal or clinical inference. Observation counts index observer effort. Host-pathogen links are observational co-occurrence."),
 ("Abbreviations", "AI avian influenza; Bd Batrachochytrium dendrobatidis; CUI UMLS Concept Unique Identifier; DOID Human Disease Ontology; EEE eastern equine encephalitis; EHR electronic health record; FIPS Federal Information Processing Standards; FRS EPA Facility Registry Service; HP Human Phenotype Ontology; HPAI highly pathogenic avian influenza; IAV influenza A virus; KG knowledge graph; MONDO Mondo Disease Ontology; NCBITaxon NCBI Taxonomy; OKN Open Knowledge Network; PFAS per- and polyfluoroalkyl substances; PFOS perfluorooctanesulfonic acid; RCM regional climate model; S2 Google S2 cell hierarchy; SDoH social determinants of health; UMLS Unified Medical Language System; WNV West Nile virus; WQP US Water Quality Portal."),
], columns=["Item", "Rule"])
sheet("Methods & Rules", methods, widths={"Item": 30, "Rule": 130}, wrap=("B",))
wb.save("Sentinel-Wildlife_results.xlsx")
print("xlsx written:", len(wb.sheetnames), "sheets")

# ------------------------------------------------------------------ HTML
rows = []
for r in c.itertuples():
    rows.append({"rank": int(r.rank), "county": r.county_name, "FIPS": r.fips5,
                 "tier": r.conf_tier, "score": round(float(r.priority), 3),
                 "sentinel_sp": int(r.sentinel_species), "proximity": r.best_tier,
                 "host_sp": int(r.host_species), "species": int(r.total_species),
                 "pfas_fac": int(r.pfas_facilities), "facilities": int(r.frs_facilities),
                 "asthma_pct": r.adult_asthma_pct,
                 "samples": 0,
                 "sources_n": int(r.sources_n),
                 "sources": ast.literal_eval(r.sources) if isinstance(r.sources, str) else [],
                 "evidence": r.evidence})
table = candidate_table(
    rows,
    columns=[("rank", "#"), ("county", "county"), ("FIPS", "FIPS"), ("tier", "tier"),
             ("score", "priority score"), ("sentinel_sp", "sentinel-capable sp."),
             ("proximity", "best proximity tier"), ("host_sp", "pathogen-host sp."),
             ("species", "species observed"), ("pfas_fac", "EPA PFAS facilities"),
             ("facilities", "EPA FRS facilities (all)"),
             ("asthma_pct", "adult asthma %"), ("samples", "contaminant samples"),
             ("evidence", "evidence supporting the rank")],
    search_keys=["county", "FIPS", "tier", "proximity", "evidence"],
    numeric_keys=["rank", "score", "sentinel_sp", "host_sp", "species", "pfas_fac", "facilities",
                  "asthma_pct", "samples"],
    page_size=25, default_sort="rank",
    extra_filters=[("tier", "confidence tier"), ("proximity", "best proximity tier")],
    sources_col=("sources_n", "sources"))

kpis = kpis_from_stats(st, [
    ("records_total", "wildlife records", ","),
    ("species_total", "taxa in the record"),
    ("fl_any_pfas_samples", "contaminant samples in Florida"),
    ("measured_species", "species with measured body burden"),
    ("inferred_species", "inferred sentinel candidates"),
    ("gap_species", "pathogen hosts never sampled"),
    ("county_tierA", "tier-A counties"),
    ("kgs_queried", "knowledge graphs queried"),
])
md = pathlib.Path("Sentinel-Wildlife_report.md")
md.write_text(fill_stats(md.read_text(), st))          # deliver a standalone .md
build_report_from_markdown(str(md), "Sentinel-Wildlife_report.html",
                           kpis=kpis, table=table, stats=st)
# splice the interactive OSM map in place of its marker (chrome, not prose)
h = pathlib.Path("Sentinel-Wildlife_report.html")
iframe = pathlib.Path("figures/_map_iframe.html").read_text()
html = h.read_text()
assert "<!-- COUNTY_MAP -->" in html, "map marker missing from rendered HTML"
h.write_text(html.replace("<!-- COUNTY_MAP -->", iframe))
from build_report_html import check_report_parity
check_report_parity(str(md), str(h))
