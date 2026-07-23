#!/usr/bin/env python3
"""Render the HTML report (from the .md) and the multi-sheet Excel workbook."""
import sys, os, json
sys.path.insert(0, "/sessions/keen-charming-hypatia/mnt/.claude/skills/okn-report-style/scripts")
import pandas as pd
import build_report_html as B

BASE="/sessions/keen-charming-hypatia/mnt/Environmental-Justice"
D=f"{BASE}/data"
stats=json.load(open(f"{D}/stats.json"))
rows=json.load(open(f"{D}/table_rows.json"))
MD=f"{BASE}/Environmental-Justice_report.md"
HTML=f"{BASE}/Environmental-Justice_report.html"

# ---- KPI cards from stats ----
kpis=B.kpis_from_stats(stats,[
    ("n_counties","U.S. counties analysed",","),
    ("n_indicators","Indicators integrated"),
    ("tierA","Tier A — very high burden"),
    ("tierB","Tier B — high burden"),
    ("pct_high","% counties high-burden",".1f"),
])

# ---- interactive results table ----
cols=[("FIPS","FIPS"),("County","County"),("State","State"),("Tier","Tier"),
      ("Consensus","Consensus"),("Env domains","Env"),("Social domains","Social"),
      ("Service scarce","Svc scarce"),("Burden index","Burden idx"),("Mismatch","Mismatch"),
      ("Facilities","Facilities"),("PFAS facilities","PFAS fac"),("sources_n","sources (n)")]
table=B.candidate_table(
    rows, cols,
    search_keys=["County","State","FIPS"],
    numeric_keys=["Consensus","Env domains","Social domains","Burden index","Mismatch","Facilities","PFAS facilities","sources_n"],
    page_size=25, default_sort="sources_n",
    extra_filters=[("Tier","Tier"),("State","State"),("Service scarce","Service scarce")],
    sources_col=("sources_n","sources_list"),
)

# Interactive county map ships as a standalone companion file (Environmental-Justice_county_map.html),
# named in the report; inlining the 153KB folium iframe stalls the parity regex, so we link not inline.
import shutil
shutil.copy(f"{D}/county_map.html", f"{BASE}/Environmental-Justice_county_map.html")
B.build_report_from_markdown(MD, HTML, kpis=kpis, table=table, stats=stats)
print("standalone interactive map: Environmental-Justice_county_map.html")

# ================= Excel workbook =================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
HDR=PatternFill("solid",fgColor="1F4E79"); HDRF=Font(name="Arial",bold=True,color="FFFFFF",size=10)
TIER={'A (very high)':"F4CCCC",'B (high)':"FCE5CD",'C (moderate)':"FFF2CC",'D (low)':"FFFFFF"}
def sheet(wb,name,df,color_tier=False,widths=None):
    ws=wb.create_sheet(name[:31])
    ws.append(list(df.columns))
    for c in range(1,len(df.columns)+1):
        cell=ws.cell(1,c); cell.fill=HDR; cell.font=HDRF; cell.alignment=Alignment(wrap_text=True,vertical="center")
    for _,r in df.iterrows():
        ws.append([None if (pd.isna(v) if not isinstance(v,(list,dict)) else False) else v for v in r.tolist()])
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
    for i,col in enumerate(df.columns,1):
        w=(widths or {}).get(col, min(max(12,int(df[col].astype(str).str.len().mean() if len(df) else 12)+2),40))
        ws.column_dimensions[get_column_letter(i)].width=w
    if color_tier and 'tier' in df.columns:
        ti=list(df.columns).index('tier')+1
        for ri in range(2,len(df)+2):
            t=ws.cell(ri,ti).value
            if t in TIER: ws.cell(ri,ti).fill=PatternFill("solid",fgColor=TIER[t])
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.font=Font(name="Arial",size=9)
    return ws

wb=openpyxl.Workbook(); wb.remove(wb.active)
mc=pd.read_csv(f"{D}/master_county.csv",dtype={'fips':str}).sort_values(['consensus','burden_index'],ascending=False)
# tidy column order for the ranked sheet
front=['fips','label','stateName','tier','consensus','env_burden','soc_burden','service_scarce','burden_index','mismatch_index','service_capacity','domains_with_data']
mc=mc[[c for c in front if c in mc.columns]+[c for c in mc.columns if c not in front]]
sheet(wb,"Ranked results",mc,color_tier=True,widths={'label':22,'stateName':16})
sheet(wb,"State summary",pd.read_csv(f"{D}/state_rollup.csv").sort_values('mean_consensus',ascending=False))
sheet(wb,"Domain correlations",pd.read_csv(f"{D}/domain_correlations.csv").rename(columns={'Unnamed: 0':'domain'}))
sheet(wb,"Indicator correlations",pd.read_csv(f"{D}/indicator_correlations.csv").rename(columns={'Unnamed: 0':'indicator'}))
sheet(wb,"PFAS ToxCast",pd.read_csv(f"{D}/mechanistic_pfas.csv"),widths={'chemical':38})
sheet(wb,"PFAS AOP-529",pd.read_csv(f"{D}/mechanistic_aop529.csv"),widths={'event':52})
ev=pd.read_csv(f"{D}/evidence_long.csv",dtype={'fips':str})
# cap the workbook sheet to burden-flagged counties (consensus>=1); full record in data/evidence_long.csv
keepf=set(mc[mc['consensus']>=1]['fips'])
evx=ev[ev['fips'].isin(keepf)]
sheet(wb,"Evidence (flagged counties)",evx,widths={'county':22,'state':16,'indicator':46,'evidence_type':22,'source_kg':18,'geo_level':20})
# Methods & rules sheet
meth=pd.DataFrame({'Item':[
  'Unit of analysis','Coverage','Integration key','Consensus score','Domain flag','Tiers','Mismatch index',
  'Evidence handling','Disease rollup','Caveat',
  'CHR','SVI','PFAS','FIPS','AOP','MIE','PPAR','ToxCast'],
  'Definition':[
  'U.S. county (5-digit FIPS)','50 states + DC (facility/PFAS/geo axes: 48 contiguous + DC)',
  'Shared county FIPS via verified OKN crosswalks (spatialkg hub)',
  'Count of 6 independent burden domains (D1-D6) in national worst quintile',
  'Domain index (mean direction-adjusted z of its indicators) >= 80th percentile',
  'A: 5-6 domains; B: 3-4; C: 1-2; D: 0',
  '(mean burden-domain z) - (service-capacity z); higher = burdened & underserved',
  'Each indicator kept separate with evidence_type/source_kg/geo_level/direction; never fused into one score',
  'CDC PLACES place-level prevalence rolled to county (population-weighted) via spoke-okn PARTOF_LpL',
  'Observational county-level associations; hypothesis-generating, not causal',
  'County Health Rankings','CDC/ATSDR Social Vulnerability Index','per- and polyfluoroalkyl substances',
  'Federal Information Processing Standards county code','Adverse Outcome Pathway','Molecular initiating event',
  'Peroxisome proliferator-activated receptor','EPA high-throughput toxicity screening']})
sheet(wb,"Methods & rules",meth,widths={'Item':22,'Definition':80})
wb.save(f"{BASE}/Environmental-Justice_results.xlsx")
print("xlsx sheets:", wb.sheetnames)
