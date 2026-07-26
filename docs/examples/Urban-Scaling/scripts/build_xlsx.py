import pandas as pd, numpy as np, json
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

res=pd.read_csv("data/ranked_results.csv")
lad=pd.read_csv("data/threshold_ladder.csv")
rob=pd.read_csv("data/scaling_exponents_robustness.csv")
band=pd.read_csv("data/band_stats_place.csv")
binp=pd.read_csv("data/binned_place_lnrate.csv")
suff=pd.read_csv("data/sufficient_stats.csv")
quad=pd.read_csv("data/quadratic_fits.csv")
ver =pd.read_csv("data/verify_diabetes_largest.csv")

methods=pd.DataFrame({"Item":[
 "Model","Estimation","Exponent meaning","Rate elasticity","R2_count (why not reported)",
 "R2_rate (reported)","Place source","County mortality source","County population",
 "Crime source","Cross-KG join","Restricted samples","Tier A","Tier B","Tier C",
 "Endpoint","Abbreviations"],
 "Definition":[
 "ln Y = ln Y0 + beta * ln N, where Y = outcome count and N = settlement population",
 "OLS from sufficient statistics (n, Sx, Sy, Sxx, Sxy, Syy) aggregated server-side in SPARQL using math:log; verified against direct OLS to 6.4e-15",
 "beta>1 superlinear (larger settlements bear disproportionately more); beta<1 sublinear; beta=1 proportional",
 "beta - 1; the elasticity of the per-capita RATE with respect to population",
 "Mechanically near 1 because counts are reconstructed as rate x N, placing N on both sides of the regression - misleading and therefore excluded",
 "R-squared of ln(Y/N) on ln(N), slope beta-1; the honest measure of how much of rate variance population explains",
 "spoke-okn PREVALENCE_DpL - CDC PLACES age-adjusted prevalence, 9 conditions, 26,343 census places, each with total_population",
 "spoke-okn PREVALENCEIN_SpL - County Health Rankings 2023, 8 mortality measures on county nodes",
 "ruralkg settlementtype:population, census 2013, joined on 5-digit county FIPS",
 "scales hasIdbCounty - 121,785 federal criminal case filings aggregated per county of origin",
 "Verified crosswalk K2: spoke-okn location IRI (FIPS5) <-> ruralkg censusCounty KWG administrativeRegion.USA.{FIPS5}; 3,196 counties",
 "Places with population >= 50,000 (n=709); metropolitan counties RUCC 2013 codes 1-3",
 "Same direction AND significant in both full and restricted samples - robust scaling behaviour",
 "Direction stable but significance changes between samples - suggestive only",
 "Direction REVERSES between samples - exponent is an artefact of the city definition",
 "OKN federated SPARQL (FRINK), QLever; queried 2026-07-25/26",
 "beta=scaling exponent; N=population; Y=outcome count; CI=confidence interval; YPLL=years of potential life lost before 75; RUCC=Rural-Urban Continuum Code; FIPS=Federal Information Processing Standards code; MRP=multilevel regression and poststratification; CHR=County Health Rankings"]})

sheets={"Ranked Results":res,"Threshold Ladder":lad,"Restricted Samples":rob,
        "Band Sufficient Stats":band,"Binned Prevalence":binp,
        "Full-Sample Sufficient Stats":suff,"Quadratic Fits":quad,
        "Verification Extract":ver,"Methods & Rules":methods}

with pd.ExcelWriter("Urban-Scaling_results.xlsx",engine="openpyxl") as xw:
    for name,df in sheets.items(): df.to_excel(xw,sheet_name=name[:31],index=False)
    wb=xw.book
    hdr=Font(name="Arial",bold=True,color="FFFFFF",size=10)
    fill=PatternFill("solid",fgColor="1F4E79")
    tier={"A":"C6EFCE","B":"FFEB9C","C":"FFC7CE"}
    for name in sheets:
        ws=wb[name[:31]]
        ws.freeze_panes="A2"
        ws.auto_filter.ref=ws.dimensions
        for c in ws[1]: c.font=hdr; c.fill=fill; c.alignment=Alignment(vertical="center",wrap_text=True)
        for col in ws.columns:
            L=get_column_letter(col[0].column)
            w=max((len(str(c.value)) for c in col if c.value is not None),default=10)
            ws.column_dimensions[L].width=min(max(w+2,11),62)
            for c in col:
                if c.row>1: c.font=Font(name="Arial",size=10)
        if name=="Ranked Results":
            ti=[c.column for c in ws[1] if c.value=="tier"][0]
            for r in range(2,ws.max_row+1):
                v=ws.cell(row=r,column=ti).value
                if v in tier:
                    for c in ws[r]: c.fill=PatternFill("solid",fgColor=tier[v])
        if name=="Methods & Rules":
            ws.column_dimensions["B"].width=110
            for r in range(2,ws.max_row+1): ws.cell(row=r,column=2).alignment=Alignment(wrap_text=True,vertical="top")
print("workbook written:", {k:len(v) for k,v in sheets.items()})
