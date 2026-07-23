import pandas as pd, numpy as np, json
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
D='/sessions/magical-intelligent-maxwell/mnt/SANS/SANS_cross_species_transcriptomics'
R=pd.read_csv(f'{D}/data/ranked_genes.csv'); M=pd.read_csv(f'{D}/data/de_master.csv')
E=pd.read_csv(f'{D}/data/enrichment_results.csv'); T=pd.read_csv(f'{D}/data/trait_enrichment.csv')
RD=pd.read_csv(f'{D}/data/rdkg_ocular_diseases.csv'); DR=pd.read_csv(f'{D}/data/prokn_core_drugs.csv')
HP=pd.read_csv(f'{D}/data/oard_hp.csv')
ranked=R[['rank','tier','humanSymbol','hEntrez','systems','n_studies','n_mouse_studies','n_fly_studies',
 'n_tissues','mouse_systems','xspecies_ok','fly_fanout','consensus_dir','dir_consistency',
 'max_abs_lfc','mouse_lfc','fly_lfc','min_padj','n_strict','studies','score']].copy()
ranked.columns=['Rank','Tier','Human symbol','Human Entrez','Systems','Studies (n)','Mouse studies','Fly studies',
 'Tissues (n)','Mouse organ systems','Cross-species (1:1)','Fly orthologue fan-out','Consensus direction',
 'Direction agreement','Max |log2FC|','Mouse log2FC','Fly log2FC','Min adj. p','Strict-threshold hits',
 'OSDR studies','Consistency score']
inv=(M.groupby(['osd','system','tissue','species']).agg(**{'DE genes':('entrez','nunique')})
       .reset_index().rename(columns={'osd':'OSDR study','system':'Organ system','tissue':'Tissue','species':'Species'}))
enr=E[['family','aspect','label','category','k','K','n','N','expected','fold','p','fdr']].copy()
enr.columns=['Family','GO aspect','Category','Category ID','Hits k','Set size K','Signature n','Background N','Expected','Fold','p','FDR']
tr=T[['trait_label','trait','k','K','expected','fold','p','fdr']].copy()
tr.columns=['Trait','Trait ID','Hits k','Set size K','Expected','Fold','p','FDR']
dis=RD[['symbol','disease_label','disease','drug_label']].drop_duplicates()
dis.columns=['Core gene','Disease','Disease ID','Drug treating this disease (rdkg)']
dr=DR[['symbol','compound_label','direction','compound']].drop_duplicates()
dr.columns=['Core gene target','Compound','Direction of modulation','Compound ID']
hp=HP[['disease_label','hp_label','hp']].drop_duplicates(); hp.columns=['Disease anchor','HPO phenotype','HPO ID']
methods=pd.DataFrame({'Item':[
 'Contrast rule','Comparability rule','DEG threshold','Ocular deviation','Ortholog projection',
 'Fan-out penalty','Consistency score','GO background','Reactome background','Trait background',
 'Statistical test','Cross-KG bridge','Level of inference','Abbreviations'],
 'Specification':[
 'Space Flight (arm 1) vs Ground Control (arm 2) only; log2FC>0 = up in spaceflight (get_valid_contrasts)',
 'Both arms share covariates after stripping condition labels/codes AND material_id_1 = material_id_2; 492 confounded assays excluded',
 'adj. p <= 0.05 AND |log2FC| >= 1',
 'Ocular arm additionally retrieved at adj. p <= 0.05 (no fold cut) because the strict cut yields only 38 rows',
 'spoke-genelab IS_ORTHOLOG_MGiG; collapsed per OSDR study by max |log2FC| (collapse_orthologs.py)',
 'Cross-species credit only when the non-mouse orthologue has fan-out <= 3 human paralogues',
 '2.5*mouse_studies + 3*tissues + 4*mouse_systems + 6*cross_species + 3.5*ocular + 2*dir_consistency*studies + 1.5*strict_hits + 2*log10(max|log2FC|+1)',
 '8290 prokn genes with >=1 GO annotation via encoded UniProt protein; n = 75 core genes mapped',
 '6032 prokn genes participating in >=1 human Reactome pathway; n = 61 core genes mapped',
 '21710 digcfdekg genes with any geneToTrait edge; n = 166 core genes mapped; 62-trait hypothesis-guided panel',
 'One-sided hypergeometric over-representation with Benjamini-Hochberg FDR; FDR < 0.05 reported',
 'spoke-genelab x spoke-okn on Entrez node-IRI (crosswalk C4); verified live = 16,326 shared gene nodes',
 'Hypothesis generation. Ortholog-inferred from model organisms; disease/drug/phenotype links are observational associations, not causal or clinical claims',
 'SANS = Spaceflight-Associated Neuro-ocular Syndrome; OSDR = NASA Open Science Data Repository; DE = differential expression; FDR = false-discovery rate; GO = Gene Ontology; BP/MF/CC = biological process / molecular function / cellular component; OXPHOS = oxidative phosphorylation; UPR = unfolded-protein response; LHON = Leber hereditary optic neuropathy; HPO = Human Phenotype Ontology']})
sheets=[('Ranked Results',ranked),('Assay Inventory',inv),('GO + Reactome Enrichment',enr),
        ('Trait Enrichment',tr),('Curated Ocular Diseases',dis),('Druggability',dr),
        ('Phenotypes (HPO)',hp),('Methods & Rules',methods)]
out=f'{D}/SANS_cross_species_transcriptomics_results.xlsx'
with pd.ExcelWriter(out,engine='openpyxl') as xw:
    for name,df in sheets: df.to_excel(xw,sheet_name=name[:31],index=False)
    wb=xw.book
    hdr=PatternFill('solid',fgColor='1F3B63'); hf=Font(name='Arial',bold=True,color='FFFFFF',size=10)
    tiers={'A':'C6EFCE','B':'FFEB9C','C':'F2F2F2'}
    for name,df in sheets:
        ws=wb[name[:31]]
        for c in ws[1]: c.fill=hdr; c.font=hf; c.alignment=Alignment(wrap_text=True,vertical='center')
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
        for i,col in enumerate(df.columns,1):
            w=min(max(12,int(df[col].astype(str).str.len().quantile(0.9) if len(df) else 12)+2),58)
            ws.column_dimensions[get_column_letter(i)].width=w
        for row in ws.iter_rows(min_row=2):
            for c in row: c.font=Font(name='Arial',size=10)
        if name=='Ranked Results':
            for row in ws.iter_rows(min_row=2,max_row=min(ws.max_row,3200)):
                t=row[1].value
                if t in tiers: row[1].fill=PatternFill('solid',fgColor=tiers[t])
        if name=='Methods & Rules':
            ws.column_dimensions['A'].width=26; ws.column_dimensions['B'].width=110
            for row in ws.iter_rows(min_row=2): row[1].alignment=Alignment(wrap_text=True,vertical='top')
print('workbook written:',out)
for n,d in sheets: print(f'  {n}: {len(d)} rows')
