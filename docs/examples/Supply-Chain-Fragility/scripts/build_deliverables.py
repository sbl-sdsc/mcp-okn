#!/usr/bin/env python3
"""Build the interactive HTML report, the interactive OSM map and the Excel workbook."""
import json, sys, pandas as pd, numpy as np
from pathlib import Path
R = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(R/"scripts"))
from build_report_html import (candidate_table, build_report_from_markdown,
                              load_stats, kpis_from_stats, fill_stats)
from okn_figstyle import folium_osm_map, folium_map_iframe
D = R/"data"
S = load_stats(str(D/"stats.json"))
SN = json.load(open(D/"stats_numeric.json"))

M   = pd.read_csv(D/"industry_master.csv", dtype={"naics":str})
lab = dict(zip(M.naics, M.label))
pc  = pd.read_csv(D/"county_burden_percapita.csv", dtype={"fips":str})
cen = pd.read_csv(D/"county_centroids.csv", dtype={"fips":str})
smm = pd.read_csv(D/"county_smm_employment.csv", dtype={"fips":str})
slc = pd.read_csv(D/"shortlist_counties.csv", dtype={"naics":str,"fips":str})

# ---------------------------------------------------------------- results table
def srcs(r):
    s = ["sudokn", "fiokg", "spatialkg"]
    if r.stackCves > 0 or r.stackPkgs > 0: s.append("securechainkg")
    if r.bridgeCves > 0: s.append("securechainkg-hw")
    return s
rows = []
for _, r in M.sort_values("jointRisk", ascending=False, na_position="last").iterrows():
    src = srcs(r)
    rows.append({
      "naics": r.naics, "industry": r.label,
      "firms": int(r.firms), "placed": int(r.placed), "jobs": int(r.empSum) if pd.notna(r.empSum) else 0,
      "facilities": int(r.facilities),
      "concIndex": round(float(r.concIndex), 2),
      "effCounties": round(float(r.effCounties), 1),
      "topCountyShare": f"{100*float(r.topCountyShare):.1f}%",
      "swScore": round(float(r.swFragility), 1),
      "stackOT": int(r.stackOT), "stackCves": int(r.stackCves),
      "bridgeCves": int(r.bridgeCves),
      "jointRisk": round(float(r.jointRisk), 1) if pd.notna(r.jointRisk) else "",
      "concTier": r.concTier, "swTier": r.swTier, "tier": r.evidenceTier,
      "shortList": "yes" if r.shortList else "no",
      "nsrc": len(src), "sources": src,
    })
tbl = candidate_table(rows,
    columns=[("naics","NAICS"),("industry","industry"),("tier","tier"),
             ("shortList","short list"),("jointRisk","joint risk"),
             ("concIndex","conc. index"),("effCounties","eff. counties"),
             ("topCountyShare","top county %"),("swScore","software score"),
             ("stackOT","OT pkgs"),("stackCves","stack CVEs"),("bridgeCves","hw CVEs"),
             ("placed","firms placed"),("jobs","recorded jobs"),("facilities","EPA facilities"),
             ("nsrc","sources (n)")],
    search_keys=["naics","industry"],
    numeric_keys=["jointRisk","concIndex","effCounties","swScore","stackOT","stackCves",
                  "bridgeCves","placed","jobs","facilities","firms"],
    default_sort="jointRisk",
    extra_filters=[("tier","evidence tier"),("concTier","concentration"),
                   ("swTier","software fragility"),("shortList","short list")],
    sources_col=("nsrc","sources"))

# ---------------------------------------------------------------- interactive map
mp = pc.merge(cen, on="fips").merge(
        smm[["fips","smmFirms","smmEmployment"]], on="fips", how="left")
mrows = []
for _, r in mp.iterrows():
    mrows.append({
      "county": r.county, "FIPS": r.fips,
      "EPA NAICS-332 facilities": int(r.n332),
      "facilities per 100k residents": round(float(r.facPer100k), 1),
      "population (ruralkg, 2013)": int(r.population),
      "RUCC class": int(r.rucc),
      "SMM firms (sudokn, all NAICS 332)": int(r.smmFirms) if pd.notna(r.smmFirms) else "not in top-45",
      "lat": float(r.lat), "lon": float(r.lng),
      "intensity": float(r.facPer100k),
    })
m = folium_osm_map(mrows, lat_key="lat", lon_key="lon", value_key="intensity",
                   tooltip_key="county", zoom_start=4, radius=4,
                   popup_keys=["county","FIPS","EPA NAICS-332 facilities",
                               "facilities per 100k residents","population (ruralkg, 2013)",
                               "RUCC class","SMM firms (sudokn, all NAICS 332)"])
map_html = folium_map_iframe(m, height=540, title="NAICS-332 intensity leaders")

# ---------------------------------------------------------------- HTML report
md_tpl = R/"data"/"Supply-Chain-Fragility_report.template.md"
md_src = R/"Supply-Chain-Fragility_report.md"
md_txt = md_tpl.read_text(encoding="utf-8")
# The folium document is ~100 kB on one line; splicing it into the Markdown before
# rendering triggers pathological regex backtracking in md_to_html, so a short sentinel
# is rendered instead and the iframe is substituted into the finished HTML.
SENTINEL = "<span id=\"okn-interactive-map\"></span>"
tmp = R/"_report_with_map.tmp.md"
tmp.write_text(md_txt.replace("<!-- INTERACTIVE_MAP -->", SENTINEL), encoding="utf-8")

kpis = kpis_from_stats(S, [
    ("smm_firms_total","SMM firms (NAICS 332)"),
    ("smm_emp_total","recorded jobs"),
    ("frs_facilities_332","EPA-regulated facilities"),
    ("sw_packages","packages in the dep. graph"),
    ("numpy_dependent_pct","% of packages depending on numpy"),
    ("shortlist_n","short-listed industries"),
    ("spearman_conc_vs_sw","concentration vs software ρ"),
    ("n_kgs","knowledge graphs joined"),
])
out_html = R/"Supply-Chain-Fragility_report.html"
build_report_from_markdown(str(tmp), out=str(out_html), kpis=kpis, table=tbl, stats=S,
                           footer="OKN federated SPARQL · mcp-okn · okn-report-style v0.1.5",
                           verify=False)
h = out_html.read_text(encoding="utf-8")
assert SENTINEL in h, "map sentinel missing from rendered HTML"
out_html.write_text(h.replace(SENTINEL, map_html), encoding="utf-8")
print("[map] interactive OSM map spliced into the HTML report")

# fill the delivered .md so it reads standalone
md_src.write_text(fill_stats(md_txt, S), encoding="utf-8")
from build_report_html import check_report_parity
check_report_parity(str(md_src), str(out_html))

# ---------------------------------------------------------------- workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
wb = Workbook(); wb.remove(wb.active)
HDR = PatternFill("solid", fgColor="3A1414")
TIER = {"A": PatternFill("solid", fgColor="D8EFD8"),
        "B": PatternFill("solid", fgColor="FDF3D8"),
        "C": PatternFill("solid", fgColor="F2F2F2")}
def sheet(name, df, tiercol=None, widths=None):
    ws = wb.create_sheet(name[:31])
    ws.append(list(df.columns))
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = HDR; c.alignment = Alignment(wrap_text=True, vertical="center")
    for rec in df.itertuples(index=False):
        ws.append(["" if (isinstance(v,float) and np.isnan(v)) else v for v in rec])
    for row in ws.iter_rows(min_row=2):
        for c in row: c.font = Font(name="Arial", size=10)
        if tiercol:
            t = row[list(df.columns).index(tiercol)].value
            if t in TIER:
                for c in row: c.fill = TIER[t]
    for i, col in enumerate(df.columns, 1):
        w = widths.get(col, 14) if widths else max(11, min(38, int(df[col].astype(str).str.len().max() or 11) + 2))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws

rt = M[["naics","label","evidenceTier","shortList","jointRisk","concIndex","effCounties",
        "topCountyFirms","topCountyShare","nCounties","hhiCounty","expHhiCounty","hhiZip3",
        "hhiState","swFragility","swTier","concTier","stackPkgs","stackOT","stackCves",
        "stackHubExposed","stackBlast","bridgeFirms","bridgeCves","firms","placed","empSum",
        "empPerFirm","facilities","facPerFirm"]].copy()
rt.columns = ["NAICS","industry","evidence tier","short list","joint risk","concentration index",
  "effective counties","top county firms","top county share","counties present","county HHI",
  "expected county HHI","ZIP3 HHI","state HHI","software score","software tier",
  "concentration tier","stack packages","OT packages","stack CVEs","stack hub-exposed",
  "stack max blast radius","bridge firms","bridge hardware CVEs","firms (national)",
  "firms placed","recorded employment","employees per firm","EPA facilities",
  "facilities per firm"]
sheet("Ranked Results", rt, tiercol="evidence tier", widths={"industry":42})

sheet("SMM by industry", pd.read_csv(D/"sudokn_industry_national.csv", dtype={"naics":str}))
sheet("EPA facilities by industry", pd.read_csv(D/"fiokg_industry_national.csv", dtype={"naics":str}))
sheet("Concentration county", pd.read_csv(D/"conc_county.csv", dtype={"naics":str}))
sheet("Concentration ZIP3", pd.read_csv(D/"conc_zip3.csv", dtype={"naics":str}))
sheet("Concentration state", pd.read_csv(D/"conc_state.csv", dtype={"naics":str}))
sheet("Counties SMM employment", smm)
sheet("Counties burden per capita", pc)
sheet("Counties burden share", pd.read_csv(D/"county_burden_share.csv", dtype={"fips":str}))
sheet("Counties burden count", pd.read_csv(D/"county_burden_count.csv", dtype={"fips":str}))
sheet("State burden vs employment", pd.read_csv(D/"state_burden_vs_employment.csv", dtype={"stateFips":str}))
sheet("Shortlist counties", slc.assign(industry=slc.naics.map(lab)))
sheet("Software fragility", pd.read_csv(D/"software_fragility_scored.csv"))
sheet("Software copyleft", pd.read_csv(D/"software_restrictive_licences.csv"))
sheet("Industrial stack", pd.read_csv(D/"industrial_stack.csv"))
sheet("Hardware CVE bridge", pd.read_csv(D/"industry_hardware_cve_bridge.csv", dtype={"naics":str}))
sheet("SDoH per capita tiers", pd.read_csv(D/"sdoh_by_intensity_percapita.csv"))
sheet("SDoH share tiers", pd.read_csv(D/"sdoh_by_intensity_share.csv"))
sheet("Firm age", pd.read_csv(D/"sudokn_firm_age.csv"))

meth = pd.DataFrame({"item": [
 "Study", "Date", "Endpoint", "Knowledge graphs (version, updated)",
 "Unit of analysis", "Universe",
 "Excluded", "County assignment",
 "Concentration index", "Sector baseline",
 "Concentration tiers", "Software fragility score",
 "Software tiers", "Joint risk", "Short list",
 "Evidence tier A", "Evidence tier B", "Evidence tier C",
 "Population tiers (per capita)", "Population tiers (share)",
 "KNOWN LIMITATION 1", "KNOWN LIMITATION 2", "KNOWN LIMITATION 3",
 "KNOWN LIMITATION 4", "KNOWN LIMITATION 5", "KNOWN LIMITATION 6",
 "Abbreviations",
], "detail": [
 "Supply-Chain-Fragility — physical manufacturing capacity x regulated burden x software dependency risk x community vulnerability",
 "2026-07-29",
 "OKN federated SPARQL via the mcp-okn MCP server",
 "sudokn v0.0.10 (2026-05-08); fiokg v0.0.11 (2026-03-18); securechainkg v0.0.11 (2026-03-23); spatialkg v0.0.6 (2026-05-07); spoke-okn v0.0.6 (2026-03-16); ruralkg v0.2.7 (2026-06-08)",
 "6-digit NAICS industry x US county (48 contiguous states + DC)",
 f"{SN['smm_industries']} six-digit NAICS 332xxx codes; {SN['smm_firms_total']:,} sudokn firms; {SN['frs_facilities_332']:,} fiokg facilities; {SN['sw_packages']:,} securechainkg packages",
 "sudokn 2- and 3-digit NAICS codes (almost entirely North Carolina — a state-specific ingest artefact that would register as spurious concentration)",
 "ZIP5 -> county crosswalk derived from fiokg: NAICS-33 facility addresses carry the ZIP inline and the facility carries county FIPS; only ZIPs where all facilities agree on one county are kept. Places 13,277 / 15,571 firms (85%)",
 "observed county HHI / expected HHI, where E[HHI] = 1/n + (1 - 1/n) * sector HHI for n placed firms",
 f"whole-sector county HHI = {SN['base_hhi_county']} ({SN['eff_counties_sector']:.0f} effective counties) over {SN['smm_firms_placed']:,} firms in {SN['smm_counties']:,} counties",
 "concentrated >= 2.0; moderate 1.5-2.0; diffuse < 1.5",
 "3.0 x (# OT/PLC-facing packages in the industry's mapped stack) + 1.5 x (recorded CVEs on those packages) + 2.0 x (# resting directly on a top-blast hub) + 8.0 if a NAICS->vendor->hardware->CVE bridge exists; rescaled to 0-100 by the maximum",
 "tertiles of the score across the 36 industries",
 "100 x percentile(concentration index) x percentile(software score), among industries with >= 60 placed firms",
 f"top {SN['shortlist_n']} by joint risk",
 "firms placed >= 100 AND concentration index >= 1.5",
 "firms placed >= 40 AND concentration index >= 1.3",
 "below either threshold, or concentration at/under the sector norm",
 "NAICS-332 facilities per 100,000 residents: high >= 15, mid 5-15, low < 5 (population from ruralkg, 2013 vintage)",
 "NAICS-332 share of the county's whole EPA-regulated facility base: dependent >= 2%, moderate 0.8-2%, marginal < 0.8%",
 "sudokn is a web-crawled directory, not a census: it records 187 Illinois fabricated-metal firms against 1,428 EPA-regulated facilities. All cross-state employment comparison is confounded",
 f"firm age is recorded for only {SN['firms_with_year']:,} firms ({SN['pct_firms_with_year']}%)",
 "NO KG edge joins the software dependency graph to a NAICS code. The industry<->software attachment is an analyst-assigned process-technology mapping (see the reproducibility record), not a graph traversal",
 f"maintainer thinness is unmeasurable: only {SN['sw_pkgs_with_contributors']:,} of {SN['sw_packages']:,} packages ({SN['pct_pkgs_with_contributors']}%) carry contributor data, on a slice disjoint from the dependency hubs",
 "blast radius is DIRECT dependency in-degree, not transitive closure (transitive reachability over 29.6M edges exceeds the endpoint budget). All blast figures are lower bounds",
 "licence data is sparse and contains at least one confirmed error (matplotlib recorded as AGPL-3.0-only). The copyleft inventory is a lower bound",
 "CVE = Common Vulnerabilities and Exposures; CWE = Common Weakness Enumeration; CHR = County Health Rankings; CMMC = Cybersecurity Maturity Model Certification; EPA FRS = EPA Facility Registry Service; FIPS = Federal Information Processing Standard; HHI = Herfindahl-Hirschman Index; ICS = industrial control system; KG = knowledge graph; MES = manufacturing execution system; NAICS = North American Industry Classification System; OPC-UA = Open Platform Communications Unified Architecture; OT = operational technology; PLC = programmable logic controller; PM2.5 = fine particulate matter; PyPI = Python Package Index; RUCC = Rural-Urban Continuum Code; S2 = Google S2 geometry cell; SDoH = social determinants of health; SMM = small and medium manufacturer; SVI = Social Vulnerability Index; ZIP5 = five-digit US postal code",
]})
sheet("Methods & Rules", meth, widths={"item":34, "detail":130})
wb.save(R/"Supply-Chain-Fragility_results.xlsx")
print("[workbook] wrote Supply-Chain-Fragility_results.xlsx with", len(wb.sheetnames), "sheets")
