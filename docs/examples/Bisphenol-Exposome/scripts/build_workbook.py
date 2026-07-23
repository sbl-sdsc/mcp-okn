#!/usr/bin/env python3
"""Build the multi-sheet Bisphenol-Exposome results workbook."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

HDR=PatternFill("solid",fgColor="1F3864"); HF=Font(name="Arial",bold=True,color="FFFFFF",size=10)
CF=Font(name="Arial",size=10); WRAP=Alignment(wrap_text=True,vertical="top")
TIER={"A":PatternFill("solid",fgColor="C6EFCE"),"B":PatternFill("solid",fgColor="FFEB9C"),"C":PatternFill("solid",fgColor="F2F2F2")}
thin=Side(style="thin",color="D9D9D9"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def sheet(wb,name,df,tiercol=None,widths=None):
    ws=wb.create_sheet(name[:31])
    for r in dataframe_to_rows(df,index=False,header=True): ws.append(r)
    for j,c in enumerate(df.columns,1):
        cell=ws.cell(1,j); cell.fill=HDR; cell.font=HF; cell.alignment=Alignment(wrap_text=True,vertical="center")
        w=(widths or {}).get(c, min(max(12,int(df[c].astype(str).str.len().mean()+6)),52))
        ws.column_dimensions[get_column_letter(j)].width=w
    for i in range(2,len(df)+2):
        for j in range(1,len(df.columns)+1):
            cell=ws.cell(i,j); cell.font=CF; cell.border=BORDER; cell.alignment=WRAP
        if tiercol and tiercol in df.columns:
            tv=str(ws.cell(i,list(df.columns).index(tiercol)+1).value)
            if tv in TIER:
                for j in range(1,len(df.columns)+1): ws.cell(i,j).fill=TIER[tv]
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
    return ws

wb=Workbook(); wb.remove(wb.active)

# 1. Ranked consensus results
cons=pd.read_csv("data/consensus_chem_disease.csv")
cons=cons.rename(columns={"genes_ICE_x_rdkg":"shared_target_genes","effect_domain":"effect_domain_match",
 "aop_curated":"curated_AOP","n_evidence_types":"n_independent_evidence","min_AC50_uM":"min_AC50_uM"})
cons=cons[["tier","chemical","disease","category","shared_target_genes","effect_domain_match","curated_AOP",
 "n_independent_evidence","min_AC50_uM","key_genes"]].sort_values(["tier","shared_target_genes"],ascending=[True,False])
sheet(wb,"Consensus chem-disease",cons,tiercol="tier",
 widths={"disease":30,"category":22,"key_genes":46,"chemical":14})

# 2. Chemical inventory
sheet(wb,"Chemical inventory",pd.read_csv("data/bisphenol_inventory.csv"),
 widths={"name":34,"primary_uses":48,"chem_class":26})

# 3. Target genes
gs=pd.read_csv("data/gene_target_summary_human.csv").sort_values(["n_bisphenols","n_assays"],ascending=False)
sheet(wb,"Target genes",gs,widths={"bisphenols":40,"symbol":12})

# 4. Reactome enrichment
rc=pd.read_csv("data/reactome_enrichment.csv").sort_values("FDR")
sheet(wb,"Reactome enrichment",rc,widths={"pwLabel":52,"genes":50})

# 5. GO enrichment
go=pd.read_csv("data/go_enrichment.csv").sort_values("FDR")
sheet(wb,"GO-BP enrichment",go,widths={"goLabel":48,"genes":50})

# 6. Disease enrichment
de=pd.read_csv("data/rdkg_disease_enrichment.csv").sort_values("FDR")
sheet(wb,"Disease enrichment",de,widths={"dlabel":42})

# 7. AOP chains + key events
aopc=pd.read_csv("data/aop_chains_summary.csv"); sheet(wb,"AOP chains",aopc,widths={"aop_title":54,"MIE":30,"adverse_outcome":30})
aopk=pd.read_csv("data/aop_key_events.csv"); sheet(wb,"AOP key events",aopk,widths={"key_event":40,"organ":16})

# 8. Effect domains
ed=pd.read_csv("data/effect_domain_by_chem.csv"); sheet(wb,"Adverse-outcome domains",ed,widths={"effect_domain":32})

# 9. Functional uses (exposure)
fu=pd.read_csv("data/ice_functional_uses.csv"); sheet(wb,"Exposure functional uses",fu,widths={"chemLabel":30,"useValue":20})

# 10. Methods & rules
methods=pd.DataFrame({"Item":[
 "Study","Unit of analysis","Chemical set","Molecular target rule","Potency","Adverse-outcome domain",
 "AOP traversal","Functional enrichment","Disease enrichment","Broad trait layer","Consensus tiers",
 "Level of inference","KGs (7)","Logged queries","",
 "ABBREVIATIONS","BPA/BPS/BPF/BPAF","BPB/BPE/BPZ/BPAP/BPP","TBBPA / TCBPA","BADGE / BisGMA",
 "AC50","AOP / MIE / KE / AO","ER/AR/PXR/PPAR","TTR / T4","FDR","MONDO / DTXSID / CAS"],
 "Definition":[
 "Bisphenol chemical exposome — federated KG map, exposure to adverse outcome",
 "Chemical-target-outcome chain for 15 actively-screened bisphenols",
 "ICE/ToxCast label+synonym match (bisphenol/sulfonyldiphenol/methylenediphenol/dihydroxydiphenyl), keyed CAS+DTXSID",
 "ICE curated-HTS endpoints with Call='Active'; Entrez target via assay_entrez_gene_id (human only)",
 "AC50 (uM), most potent per target",
 "ICE mayInformOn annotation (Cancer, DART, CardioTox, Estrogen, Androgen, Thyroid Hormone, ...)",
 "chemical<-has_chemical_entity<-stressor<-NCIT_C54571<-AOP; then has_molecular_initiating_event / has_key_event / has_adverse_outcome",
 "prokn symbol->UniProt->GO(RO_0002331)/Reactome(RO_0000056); hypergeometric + BH FDR; GO-BP N=7663, Reactome N=6032; GO MF/CC skipped (BP answers 'which programs')",
 "rdkg gene biolink:related_to MONDO; hypergeometric + BH FDR; background N=9080; 232 diseases (>=6 targets) tested",
 "digcfdekg geneToTrait (GWAS) - descriptive, broad-by-construction, kept separate",
 "A=3 independent evidence types OR >=20 shared genes+effect domain; B=2 types; C=1 type",
 "Observational / hypothesis-generating; NOT causal or clinical",
 "biobricks-ice, biobricks-toxcast, biobricks-aopwiki, spoke-okn, rdkg, prokn, digcfdekg",
 "16 non-exploratory SPARQL queries (see reproducibility record)","",
 "","bisphenol A / S / F / AF","bisphenol B / E / Z / AP / P","tetra-bromo / tetra-chloro bisphenol A",
 "bisphenol A diglycidyl ether / glycidyl methacrylate","half-maximal activity concentration (uM)",
 "adverse outcome pathway / molecular initiating event / key event / adverse outcome",
 "estrogen / androgen / pregnane-X / peroxisome-proliferator-activated receptor",
 "transthyretin / thyroxine","false-discovery rate (Benjamini-Hochberg)",
 "Mondo disease id / EPA DSSTox substance id / CAS registry number"]})
sheet(wb,"Methods & Rules",methods,widths={"Item":28,"Definition":92})

wb.save("Bisphenol-Exposome_results.xlsx")
print("workbook saved:",len(wb.sheetnames),"sheets:",wb.sheetnames)
