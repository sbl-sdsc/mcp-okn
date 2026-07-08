import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
D="data"; OUT="/sessions/affectionate-jolly-ramanujan/mnt/bone-health/bone_spaceflight_candidates.xlsx"
c=pd.read_csv(f"{D}/RANKED_bone_candidates.tsv",sep="\t")
# tidy columns
c=c.rename(columns={'humanSymbol':'human_gene','symbol':'mouse_gene','l2fc_WT':'log2FC_WT','p_WT':'adjP_WT',
    'l2fc_KO':'log2FC_Nrf2KO','p_KO':'adjP_Nrf2KO','n_arms':'sig_arms','both_samedir':'both_same_dir',
    'nrf2_dependent':'Nrf2_dependent','bone_role':'canonical_bone_role','bone_evidence':'bone_evidence',
    'digcfde_bonecat':'GWAS_bone_traits','rdkg_pheno':'HPO_bone_phenotype','nTissues':'nOtherTissues_sig'})
cols=['human_gene','mouse_gene','hEntrez','direction','log2FC_WT','adjP_WT','log2FC_Nrf2KO','adjP_Nrf2KO',
      'sig_arms','both_same_dir','Nrf2_dependent','canonical_bone_role','HPO_bone_phenotype','GWAS_bone_traits',
      'osteoarthritis','specificity','nOtherTissues_sig','priority','tier']
c=c[cols].round(3)

HEAD=PatternFill('solid',fgColor='1F3864'); HF=Font(name='Arial',bold=True,color='FFFFFF',size=10)
AF=Font(name='Arial',size=10); thin=Side(style='thin',color='D9D9D9'); border=Border(bottom=thin)
tiercol={'A':'C6EFCE','B':'FFEB9C','C':'F2F2F2'}
def style_header(ws,ncol):
    for j in range(1,ncol+1):
        cell=ws.cell(1,j); cell.fill=HEAD; cell.font=HF; cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.freeze_panes='A2'; ws.row_dimensions[1].height=30

wb=Workbook(); ws=wb.active; ws.title='Ranked Candidates'
ws.append(cols)
for _,r in c.iterrows(): ws.append(list(r.values))
style_header(ws,len(cols))
for i in range(2,ws.max_row+1):
    t=ws.cell(i,len(cols)).value
    fill=PatternFill('solid',fgColor=tiercol.get(t,'FFFFFF'))
    for j in range(1,len(cols)+1):
        cell=ws.cell(i,j); cell.font=AF; cell.border=border
        if j==len(cols): cell.fill=fill; cell.alignment=Alignment(horizontal='center')
widths=[11,11,9,8,9,9,11,11,7,9,10,22,24,22,8,13,10,8,5]
for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}1"

# Cohort sheet
ws2=wb.create_sheet('Cohort')
coh=[['OSD study','Tissue','Condition','Clean contrasts','Sig DE genes (adjp<=0.05)','Human orthologs','Role'],
 ['OSD-690','bone marrow','Space Flight vs Ground, Wild-Type',1,3161,3112,'PRIMARY flight signature'],
 ['OSD-690','bone marrow','Space Flight vs Ground, Nrf2-KO',1,3517,3537,'Oxidative-stress genetic control'],
 ['OSD-467','cortical bone','Hindlimb Unloaded vs Loaded (ground)',1,8,6,'Only mineralized-bone data (sparse)'],
 ['OSD-214','bone marrow','Hindlimb Unloaded vs Loaded (ground, +/- immunization)',4,2,0,'Near-zero-count artefacts'],]
for r in coh: ws2.append(r)
style_header(ws2,7)
for i in range(2,ws2.max_row+1):
    for j in range(1,8): ws2.cell(i,j).font=AF
for j,w in enumerate([10,13,42,15,22,15,32],1): ws2.column_dimensions[get_column_letter(j)].width=w

# Nrf2-dependent set
ws3=wb.create_sheet('Nrf2-dependent set')
nrf=[['human_gene','log2FC_Nrf2KO','adjP_Nrf2KO','bone_role','note'],
 ['COL1A1',-1.347,0.050,'type I collagen (bone matrix)','significant only in Nrf2-KO arm (trend)'],
 ['MMP9',-0.672,0.038,'osteoclast matrix metalloproteinase','significant only in Nrf2-KO arm'],
 ['NFATC1',-0.598,0.028,'master osteoclast transcription factor','significant only in Nrf2-KO arm'],
 ['LRP5',-0.511,0.046,'Wnt co-receptor / bone-mass gene','significant only in Nrf2-KO arm'],
 ['CSF1',-0.379,0.048,'M-CSF, osteoclast differentiation','significant only in Nrf2-KO arm'],
 ['LRP4',-0.331,0.082,'Wnt co-receptor / sclerostin partner','trend only'],]
for r in nrf: ws3.append(r)
style_header(ws3,5)
for i in range(2,ws3.max_row+1):
    for j in range(1,6): ws3.cell(i,j).font=AF
for j,w in enumerate([12,15,13,38,40],1): ws3.column_dimensions[get_column_letter(j)].width=w

# Enrichment
ws4=wb.create_sheet('Bone-loss enrichment')
en=[['Bone-loss gene set','Set size (K)','Background (N)','Signature in bg (n)','Observed (k)','Expected','Fold','Hypergeometric p'],
 ['digcfdekg GWAS (BMD/osteoporosis/fracture)',3412,21052,3021,492,489.6,'1.00x',0.459],
 ['rdkg curated Mendelian (HPO Osteoporosis/Osteopenia/Reduced-BMD/fracture)',148,21052,3021,31,21.2,'1.46x',0.018],]
for r in en: ws4.append(r)
style_header(ws4,8)
for i in range(2,ws4.max_row+1):
    for j in range(1,9): ws4.cell(i,j).font=AF
for j,w in enumerate([48,13,15,18,12,10,7,16],1): ws4.column_dimensions[get_column_letter(j)].width=w

# Methods
ws5=wb.create_sheet('Methods & Rules')
meth=[['Item','Value'],
 ['Endpoint','OKN federated SPARQL (https://apps.okn.us/federation/sparql)'],
 ['Primary KG','spoke-genelab v0.0.2 (2026-03-13)'],
 ['Context KGs','digcfdekg v0.0.1 · rdkg v0.0.1 · spoke-okn v0.0.6 · biobricks-aopwiki v0.0.4 · prokn v0.0.5'],
 ['Direction rule','factor_space_1="Space Flight" AND factor_space_2="Ground Control"; log2FC>0 = up in flight'],
 ['Comparability','arms matched on all covariates (genotype-clean WT and Nrf2-KO contrasts)'],
 ['Significance','adj_p_value <= 0.05 (primary); |log2FC| >= 1 effect-size cut; |log2FC|>=10 flagged artefact'],
 ['Ortholog collapsing','IS_ORTHOLOG_MGiG; max|log2FC| for 1:many/many:1 with ambiguity flag'],
 ['Cross-KG joins','Entrez node-IRI (direct): spoke-okn 16,326 · digcfdekg 19,747 · rdkg 9,034 (identifiers.org)'],
 ['Reproducibility axis','cross-genotype: WT vs Nrf2-KO within OSD-690 (1,754 both-significant, 98.4% same direction)'],
 ['Framing','mouse-derived, ortholog-inferred; hypothesis generation, not clinical inference'],
 ['Abbreviations','BMD = bone mineral density; DE = differentially expressed gene; HLU = hindlimb unloading; SF-vs-GC = Space-Flight-vs-Ground-Control; GO = Gene Ontology; FDR = false-discovery rate; GWAS BMD = bone-mineral-density GWAS gene (digcfdekg)'],]
for r in meth: ws5.append(r)
style_header(ws5,2)
for i in range(2,ws5.max_row+1):
    for j in range(1,3): ws5.cell(i,j).font=AF; ws5.cell(i,j).alignment=Alignment(wrap_text=True,vertical='top')
ws5.column_dimensions['A'].width=22; ws5.column_dimensions['B'].width=90

# GO enrichment sheet (prokn, bridged)
import pandas as _pd
go=_pd.read_csv('data/prokn_go_enrichment_labeled.tsv',sep='\t').sort_values('p')
go=go[['go','label','theme','K','k','exp','fold','p','fdr']].rename(columns={'go':'GO_id','label':'GO_term','K':'term_genes','k':'signature_genes','exp':'expected','p':'p_value','fdr':'FDR'})
go['p_value']=go['p_value'].map(lambda x:f"{x:.2e}"); go['FDR']=go['FDR'].map(lambda x:f"{x:.2e}")
wsg=wb.create_sheet('GO enrichment (prokn)')
wsg.append(list(go.columns))
for _,r in go.iterrows(): wsg.append(list(r.values))
style_header(wsg,len(go.columns))
for i in range(2,wsg.max_row+1):
    for j in range(1,len(go.columns)+1): wsg.cell(i,j).font=AF
for j,w in enumerate([12,42,30,11,14,10,7,11,11],1): wsg.column_dimensions[get_column_letter(j)].width=w
wsg.auto_filter.ref=f"A1:{get_column_letter(len(go.columns))}1"

# Countermeasures sheet
cm=_pd.read_csv('data/countermeasures.tsv',sep='\t')
wsc=wb.create_sheet('Countermeasures')
hdrs=['Countermeasure','Target axis / mechanism','Supporting signature genes','Supporting GO / pathway','Example agents','Confidence','Rationale']
wsc.append(hdrs)
for _,r in cm.iterrows(): wsc.append([r.countermeasure,r.target_axis,r.supporting_signature_genes,r.supporting_GO_pathway,r.example_agents,r.confidence,r.rationale])
style_header(wsc,len(hdrs))
for i in range(2,wsc.max_row+1):
    for j in range(1,len(hdrs)+1): wsc.cell(i,j).font=AF; wsc.cell(i,j).alignment=Alignment(wrap_text=True,vertical='top')
for j,w in enumerate([26,26,34,32,30,15,44],1): wsc.column_dimensions[get_column_letter(j)].width=w
wsc.insert_rows(1); wsc['A1']='Mechanism-derived research hypotheses - NOT medical advice; mouse-derived, ortholog-inferred.'
wsc['A1'].font=Font(name='Arial',bold=True,italic=True,color='C0392B')


# Reactome enrichment sheet
rx=_pd.read_csv('data/prokn_reactome_enrichment_labeled.tsv',sep='\t').sort_values('p')
rx=rx[['reactome','pathway','theme','K','k','exp','fold','p','fdr']].rename(columns={'reactome':'Reactome_id','pathway':'pathway','K':'pathway_genes','k':'signature_genes','exp':'expected','p':'p_value','fdr':'FDR'})
rx['p_value']=rx['p_value'].map(lambda x:f"{x:.2e}"); rx['FDR']=rx['FDR'].map(lambda x:f"{x:.2e}")
wsx=wb.create_sheet('Reactome enrichment')
wsx.append(list(rx.columns))
for _,r in rx.iterrows(): wsx.append(list(r.values))
style_header(wsx,len(rx.columns))
for i in range(2,wsx.max_row+1):
    for j in range(1,len(rx.columns)+1): wsx.cell(i,j).font=AF
for j,w in enumerate([14,52,26,13,14,10,7,11,11],1): wsx.column_dimensions[get_column_letter(j)].width=w
wsx.auto_filter.ref=f"A1:{get_column_letter(len(rx.columns))}1"

# Retrieved drugs sheet
dg=_pd.read_csv('data/retrieved_drugs.tsv',sep='\t')
wsd=wb.create_sheet('Retrieved drugs')
hd=['Drug class','Example agents (retrieved, rdkg)','Treats (rdkg bone disease)','Linked signature genes / axis','Source']
wsd.append(hd)
for _,r in dg.iterrows(): wsd.append([r.drug_class,r.example_agents_retrieved,r.treats_bone_disease_rdkg,r.linked_signature_genes_axis,r.source])
style_header(wsd,len(hd))
for i in range(2,wsd.max_row+1):
    for j in range(1,len(hd)+1): wsd.cell(i,j).font=AF; wsd.cell(i,j).alignment=Alignment(wrap_text=True,vertical='top')
for j,w in enumerate([26,44,40,40,18],1): wsd.column_dimensions[get_column_letter(j)].width=w
wsd.insert_rows(1); wsd['A1']='Curated from rdkg `treats` (Drug->bone disease). Human-disease drugs, NOT spaceflight-validated; not medical advice.'
wsd['A1'].font=Font(name='Arial',bold=True,italic=True,color='C0392B')

wb.save(OUT)
print("added Reactome + Retrieved-drugs sheets")

print("added GO + Countermeasures sheets")

print("saved",OUT,"| candidate rows:",len(c))
