import sys, json, warnings; warnings.filterwarnings("ignore")
sys.path.insert(0,"/sessions/focused-magical-ramanujan/mnt/outputs/Alzheimers_OKN/scripts")
sys.path.insert(0,"/sessions/focused-magical-ramanujan/mnt/outputs/Alzheimers_OKN/data")
import pandas as pd, numpy as np
import rdkg_ad_layers as R, prokn_drug_targets as P
D="/sessions/focused-magical-ramanujan/mnt/outputs/Alzheimers_OKN/data/"
rank=pd.read_csv(D+"ad_ranked_genes.csv"); go=pd.read_csv(D+"enrichment_GO.csv")
rx=pd.read_csv(D+"enrichment_Reactome.csv"); de=pd.read_csv(D+"gxa_ad_de.csv")
cons=pd.read_csv(D+"gxa_ad_de_consensus.csv"); bm=pd.read_csv(D+"biomarkerkg_ad_raw.csv")
oard=pd.read_csv(D+"oard_ad_assoc_ranked.csv"); bh=pd.read_csv(D+"biohealth_ad_risk_protective.csv")
pv=pd.read_csv(D+"spoke_ad_prevalence_2019.csv"); dig=pd.read_csv(D+"digcfdekg_ad_gene_trait.csv")
tiers=pd.read_csv(D+"prokn_ad_drug_tiers.csv"); pvar=pd.read_csv(D+"prokn_ad_variants.csv")

stats = dict(
  n_kgs_queried=8, n_mondo_terms=22, n_xrefs=235,
  n_genes_total=int(len(rank)), n_tierA=int((rank.tier=="A").sum()),
  n_tierB=int((rank.tier=="B").sum()), n_tierC=int((rank.tier=="C").sum()),
  n_multi_source=int((rank.n_sources>=2).sum()), n_4source=int((rank.n_sources>=4).sum()),
  n_5source=int((rank.n_sources>=5).sum()),
  n_ncrna=int((rank.biotype=="non-coding RNA").sum()),
  n_dig_assoc=int(len(dig)), n_dig_traits=int(dig.traitLabel.nunique()), n_dig_genes=int(dig.sym.nunique()),
  n_prokn_variants=int(pvar.variant.nunique()), n_prokn_var_genes=int(pvar.sym.nunique()),
  n_go_tested=int(len(go)), n_go_sig=int((go.fdr<0.05).sum()), go_N=8290, go_n=189,
  n_rx_tested=int(len(rx)), n_rx_sig=int((rx.fdr<0.05).sum()), rx_N=6032, rx_n=156,
  n_de_records=int(len(de)), n_de_genes=int(de.sym.nunique()), n_de_regions=int(de.region.nunique()),
  n_de_5region=int((cons.consistent>=5).sum()), n_de_3region=int((cons.consistent>=3).sum()),
  n_de_conflict=int(cons.conflict.sum()),
  n_drugs=int(len(tiers)), n_drugs_approved=int(tiers.approval.notna().sum()),
  n_drugs_investigational=int(tiers.approval.isna().sum()),
  n_drug_target_pairs=len(P.PAIRS), n_drug_targets=len(set(t for _,t in P.PAIRS)),
  n_contraindicated=len(R.CONTRA), n_exposures=len(R.EXPOSURES),
  n_biomarkers=int(bm.bLabel.nunique()), n_bm_diag=int((bm.rel=="diagnostic_for").sum()),
  n_bm_prog=int((bm.rel=="prognostic_for").sum()), n_bm_risk=int((bm.rel=="indicates_risk_of_developing").sum()),
  n_oard_assoc=689, n_oard_hp=620, n_oard_mondo=69,
  n_bh_predispose=int((bh.rel=="predisposes_to_condition").sum()),
  n_bh_prevent=int((bh.rel=="preventative_for_condition").sum()),
  n_bh_neg_pred=int((bh.rel=="NEG_PREDISPOSES").sum()), n_bh_neg_prev=int((bh.rel=="NEG_PREVENTS").sum()),
  n_conflict_both=219, n_conflict_pred=115, n_conflict_prev=26,
  n_countries=int(len(pv)), prev_median=round(float(pv.mid.median()),2),
  prev_max=round(float(pv.mid.max()),2), prev_max_country=str(pv.iloc[0].label),
  prev_min=round(float(pv.mid.min()),3), prev_min_country=str(pv.iloc[-1].label),
  prev_ratio=int(round(pv.mid.max()/pv.mid.min())),
  top_go=str(go.iloc[0].goLabel), top_go_fold=round(float(go.iloc[0].fold),1),
  top_rx=str(rx.iloc[0].pwLabel), top_rx_fold=round(float(rx.iloc[0].fold),1),
  n_rdkg_phen=len(R.PHENOTYPES), n_rdkg_genes=len(R.GENES), n_rdkg_ncrna=len(R.NCRNA),
)
json.dump(stats, open(D+"stats.json","w"), indent=1)
print(json.dumps(stats, indent=1)[:1400])

# ---------------- workbook ----------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
TIERC={"A":"FFF4CCCC","B":"FFFCE8CC","C":"FFEFEFEF"}
def sheet(wb, name, df, tiercol=None, widths=None):
    ws=wb.create_sheet(name[:31])
    for r in dataframe_to_rows(df, index=False, header=True): ws.append(r)
    for c in ws[1]:
        c.font=Font(name="Arial", bold=True, color="FFFFFFFF"); c.fill=PatternFill("solid", fgColor="FF37474F")
        c.alignment=Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for i,col in enumerate(df.columns,1):
        w = (widths or {}).get(col, min(38, max(11, int(df[col].astype(str).str.len().quantile(0.92))+3)))
        ws.column_dimensions[get_column_letter(i)].width=w
    if tiercol:
        ti=list(df.columns).index(tiercol)+1
        for row in ws.iter_rows(min_row=2):
            t=row[ti-1].value
            if t in TIERC:
                for c in row: c.fill=PatternFill("solid", fgColor=TIERC[t])
    for row in ws.iter_rows(min_row=2):
        for c in row: c.font=Font(name="Arial", size=10)
    return ws

wb=Workbook(); wb.remove(wb.active)
sheet(wb,"Ranked Results", rank, tiercol="tier", widths={"gene":12,"sources":40,"evidence_types":34,"drugs":38,"secondary_support":34})
sheet(wb,"GO enrichment", go[["goLabel","ns","k","K","n","N","expected","fold","p","fdr"]], widths={"goLabel":52})
sheet(wb,"Reactome enrichment", rx[["pwLabel","k","K","n","N","expected","fold","p","fdr"]], widths={"pwLabel":58})
sheet(wb,"Differential expression", de[["region","assayName","sym","dir","log2fc","adjp"]], widths={"assayName":52})
sheet(wb,"DE regional consensus", cons, widths={"sym":14})
sheet(wb,"digcfdekg gene-trait", pd.read_csv(D+"digcfdekg_ad_gene_trait.csv"), widths={"traitLabel":46})
sheet(wb,"prokn curated variants", pvar[["dLabel","sym","prot","protLabel","variant"]], widths={"protLabel":46,"variant":54})
sheet(wb,"Therapeutics", tiers, widths={"label":34,"tier":30})
sheet(wb,"Drug-target pairs", pd.DataFrame(P.PAIRS, columns=["drug","target_gene"]), widths={"drug":30})
sheet(wb,"Contraindicated drugs", pd.DataFrame({"drug":R.CONTRA,"source":"rdkg biolink:contraindicated_for"}))
sheet(wb,"Biomarkers (diag+prog)", pd.read_csv(D+"biomarkerkg_ad_diag_prog.csv")[["rel","bLabel","aLabel","sLabel"]], widths={"bLabel":60})
sheet(wb,"Biomarkers (all)", bm[["rel","bLabel","aLabel","sLabel"]], widths={"bLabel":66})
sheet(wb,"Phenotypes (oard-kg)", oard, widths={"otherLabel":46})
sheet(wb,"Risk & protective", bh[["rel","otherLabel","cat","source","nStatements"]], widths={"otherLabel":44})
sheet(wb,"Environmental exposures", pd.DataFrame({"exposure":R.EXPOSURES,"predicate":"rdkg biolink:contributes_to"}))
sheet(wb,"Prevalence by country", pv, widths={"label":30})
meth=pd.DataFrame({"item":[
 "Study","Anchor term","Subtype closure","KGs queried","Gene sets (primary)","Gene sets (secondary)",
 "Consensus rule","Tier A","Tier B","Tier C","Composite score",
 "GO enrichment","Reactome enrichment","Enrichment background (GO)","Enrichment background (Reactome)",
 "DE thresholds","Evidence types (kept separate)","Level of inference","Abbreviations"],
 "value":[
 "Alzheimer's disease multi-knowledge-graph integrative map (OKN federated SPARQL)",
 "MONDO:0004975 (Alzheimer disease)",
 "ubergraph rdfs:subClassOf* -> 22 MONDO terms; 235 identifier cross-references",
 "spoke-okn, digcfdekg, prokn, rdkg, biomarkerkg, gene-expression-atlas-okn, oard-kg, biohealth (+ ubergraph as ontology bridge)",
 "spoke-okn ASSOCIATES_DaG; digcfdekg geneToTrait (7 core AD traits); prokn UniProt natural-variant annotations; biomarkerkg dbSNP risk markers; gene-expression-atlas DE in >=3 brain regions; rdkg related_to",
 "digcfdekg family-history proxy traits; digcfdekg CSF/imaging endophenotypes; gene-expression-atlas DE in >=2 regions",
 "gene counted once per PRIMARY source; sources are independent KGs, not independent studies",
 ">=3 primary KG sources AND >=2 distinct evidence types",
 ">=2 primary KG sources","1 primary KG source",
 "3*n_sources + 2*n_evidence_types + 1*n_secondary + 2*druggable + 0.75*min(variants,10) + 1*DE_regions + 0.75*min(PIGEAN,6)",
 "hypergeometric + Benjamini-Hochberg FDR; k>=3, K>=3",
 "hypergeometric + Benjamini-Hochberg FDR; k>=3, K>=3",
 "8,290 prokn genes with >=1 GO annotation (explicit, not whole-genome)",
 "6,032 prokn genes with >=1 human Reactome (R-HSA) pathway",
 "Gene Expression Atlas adjusted p <= 0.05 as loaded; direction as recorded by the source",
 "curated knowledge | genetic association | differential molecular activity | pathway membership | clinical association | literature-derived",
 "HYPOTHESIS GENERATION. Observational and curatorial associations, not causal or clinical inference.",
 "AD Alzheimer's disease; Ab amyloid-beta; APP amyloid precursor protein; BBB blood-brain barrier; CAA cerebral amyloid angiopathy; CI confidence interval; CL Cell Ontology; DE differential expression; DOID Disease Ontology; EFO Experimental Factor Ontology; EOAD early-onset AD; FDR false-discovery rate; GO Gene Ontology; GWAS genome-wide association study; HP Human Phenotype Ontology; KG knowledge graph; LOAD late-onset AD; LUBAC linear ubiquitin chain assembly complex; MCI mild cognitive impairment; MONDO Mondo Disease Ontology; NFT neurofibrillary tangle; ORA over-representation analysis; PIGEAN CFDE REVEAL gene-trait scoring method; SDoH social determinants of health; UBERON Uber-anatomy ontology; UMLS Unified Medical Language System"]})
sheet(wb,"Methods & Rules", meth, widths={"item":30,"value":110})
out="/sessions/focused-magical-ramanujan/mnt/outputs/Alzheimers_OKN/Alzheimers_results.xlsx"
wb.save(out); print("\nworkbook:", out, "| sheets:", len(wb.sheetnames))
