#!/usr/bin/env python3
"""Instrument-Criticality — multi-sheet results workbook."""
import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUT = ROOT / "Instrument-Criticality_results.xlsx"

sci = pd.read_csv(DATA / "instrument_criticality.csv")
full = pd.read_csv(DATA / "instrument_catalogue_full.csv")
stats = json.loads((DATA / "stats.json").read_text())
rc = json.loads((DATA / "risk_classes.json").read_text())
ragree = pd.read_csv(DATA / "route_agreement_matrix.csv", index_col=0)
geo = pd.read_csv(DATA / "geo_paper_country_mentions.csv")
reg = pd.read_csv(DATA / "geo_study_regions.csv")
coh = pd.read_csv(DATA / "cohort_institution_countries.csv")
r5 = pd.read_csv(DATA / "sole_measured_variables_resolved.csv")

ranked = sci.rename(columns={
    "instr": "instrument", "k": "match key", "criticality": "criticality score",
    "footprint_score": "footprint score (raw datasets)",
    "footprint_frac_score": "footprint score (fractional)",
    "rank_gap": "footprint rank - criticality rank",
    "corroboration": "corroborating routes (0-5)",
    "DB": "dependency breadth (0-1)", "IR": "irreplaceability (0-1)",
    "nPlat": "platforms", "nDs": "datasets (raw attribution)",
    "fracDs": "datasets (fractional attribution)",
    "allPubs": "publications citing its datasets",
    "nKeywords": "GCMD science keywords", "nProjects": "projects",
    "nDaacs": "data centres", "firstStartYear": "earliest dataset start",
    "lastStartYear": "latest dataset start",
    "R1_cmMentionPapers": "R1 papers naming instrument",
    "R1b_platMentionPapers": "R1b papers naming platform",
    "R2_cmModelPapers": "R2 papers naming instrument + using model",
    "R2_cmDistinctModels": "R2 distinct models",
    "R3_doiPapers": "R3 DOI-matched papers citing its datasets",
    "R4_modelTitlePubs": "R4 NASA modelling-title papers",
    "R5_nModelVars": "R5 model-produced variables measured",
    "R5_nSoleVars": "R5 variables solely measured",
    "R5_soleVars": "R5 sole variable list",
    "soleKeyword": "sole-source GCMD keyword",
    "riskClass": "risk class", "tier": "tier", "category": "label category",
    "nSources": "sources (n)"})
ranked["sources"] = ["nasa-gesdisc-kg; climatemodelskg" if n == 2
                     else "nasa-gesdisc-kg" for n in ranked["sources (n)"]]

methods = pd.DataFrame([
    ("Study", "Instrument-Criticality — what climate modelling would lose if an "
     "Earth-observation instrument went dark"),
    ("Endpoint", "OKN federated SPARQL (Proto-OKN)"),
    ("Knowledge graphs", "nasa-gesdisc-kg v0.0.6 (updated 2026-06-08); "
     "climatemodelskg v0.0.15 (updated 2026-05-06)"),
    ("Unit of analysis", "GCMD instrument label carried by a spaceborne platform "
     "in nasa-gesdisc-kg"),
    ("Scoping rule 1", "Spaceborne only: platform dc:type in {Earth Observation "
     "Satellites, Space Stations/Crewed Spacecraft, Solar/Space Observation "
     "Satellites, Navigation Satellites, Spacecraft, Space-based Platforms} "
     "→ 254 platforms, 288 instrument labels, 4,931 datasets"),
    ("Scoping rule 2", "288 labels hand-classified into 243 science instruments, "
     "25 generic GCMD class labels (e.g. RADIOMETERS, SAR, GPS, NOT APPLICABLE) "
     "and 20 platform/bus subsystems (e.g. Star Tracker, GYROS, GRACE SCA). "
     "Only the 243 science instruments are scored."),
    ("Attribution rule", "nasa-gesdisc-kg has NO Dataset→Instrument edge. Datasets "
     "attach to PLATFORMS (HAS_PLATFORM) and platforms carry instruments "
     "(HAS_INSTRUMENT), so every instrument on a platform inherits all of that "
     "platform's datasets. Raw and fractional (datasets ÷ instruments on the "
     "platform) footprints are both reported."),
    ("Route R1 (textual)", "climatemodelskg Paper -PAPER_MENTIONS-> Instrument, "
     "instrument name matched case-insensitively to the GCMD label. NLP-extracted "
     "from paper text — textual evidence."),
    ("Route R1b (textual)", "climatemodelskg Paper -PAPER_MENTIONS-> Platform, "
     "platform name matched to the GCMD platform label, then platform "
     "-HAS_INSTRUMENT-> instrument. Textual, coarser than R1."),
    ("Route R2 (hybrid)", "R1 restricted to papers that ALSO carry "
     "PAPER_USES_MODEL -> Source. Textual mention x structural model link."),
    ("Route R3 (structural)", "climatemodelskg Paper doi == nasa-gesdisc-kg "
     "Publication bibo:doi (bare-DOI normalisation), then Publication "
     "-USES_DATASET-> Dataset -HAS_PLATFORM-> Platform -HAS_INSTRUMENT-> "
     "Instrument. NASA's own record that a climate-modelling paper used that "
     "instrument's data."),
    ("Route R4 (nasa-internal)", "nasa-gesdisc-kg Publication with a modelling "
     "term in schema:title (climate model | cmip | earth system model | "
     "reanalysis | general circulation model) AND -USES_DATASET->. Independent of "
     "climatemodelskg entirely. Textual title filter, structural dataset link."),
    ("Route R5 (capability)", "climatemodelskg Instrument -MEASURES_VARIABLE-> "
     "Variable <-PRODUCES_VARIABLE- SourceComponent. 184 of 3,144 variables are "
     "both measured and model-produced; an instrument is a SOLE MEASURER where no "
     "other distinct instrument name measures that variable."),
    ("Dependency breadth (DB)", "unweighted mean of log1p-then-min-max-normalised "
     "R1, R1b, R2, R3, R4"),
    ("Irreplaceability (IR)", "log1p-min-max of (R5 sole-measured variables + 1 if "
     "the instrument is the only spaceborne source of a GCMD science keyword)"),
    ("Criticality score", "100 x (0.55*DB + 0.30*IR + 0.15*(corroboration/5)) "
     "rescaled so the maximum = 100"),
    ("Tier A", "corroboration >= 4 routes AND criticality >= 30"),
    ("Tier B", "corroboration >= 2 routes"),
    ("Tier C", "corroboration <= 1 route"),
    ("Risk class A", "DB >= 75th percentile AND corroboration >= 3"),
    ("Risk class B", "irreplaceability > 0 AND criticality < 75th percentile"),
    ("Risk class C", "datasets >= median AND all five dependency routes = 0"),
    ("Asymmetry test", "Spearman rank correlation between data footprint and "
     "criticality, plus rank-gap (footprint rank - criticality rank)"),
    ("Level of inference", "Bibliometric and catalogue-structural association. NOT "
     "a measurement of scientific irreplaceability, and NOT an engineering or "
     "programmatic risk assessment."),
    ("Abbreviations", "GCMD = Global Change Master Directory; DAAC = Distributed "
     "Active Archive Center; CMR = Common Metadata Repository; DOI = Digital "
     "Object Identifier; ORCID = Open Researcher and Contributor ID; ROR = "
     "Research Organization Registry; CMIP = Coupled Model Intercomparison "
     "Project; ECV = Essential Climate Variable; GCOS = Global Climate Observing "
     "System; ERB = Earth Radiation Budget; NLP = natural-language processing; "
     "KG = knowledge graph; DB = dependency breadth; IR = irreplaceability; "
     "rho = Spearman rank correlation coefficient"),
], columns=["Item", "Definition"])

kv = pd.DataFrame(sorted(
    (k, json.dumps(v) if isinstance(v, list) else v) for k, v in stats.items()),
    columns=["statistic", "value"])

sheets = {
    "Ranked Results": ranked.drop(columns=["sources (n)"]).assign(
        **{"sources (n)": sci["nSources"]}),
    "Catalogue Inventory": full[[
        "instr", "k", "category", "nPlat", "nDs", "fracDs", "allPubs",
        "nKeywords", "nProjects", "nDaacs", "firstStartYear",
        "lastStartYear"]],
    "Risk Class A": pd.DataFrame(rc["A"]),
    "Risk Class B": pd.DataFrame(rc["B"]),
    "Risk Class C": pd.DataFrame(rc["C"]),
    "Route Agreement": ragree.reset_index().rename(columns={"index": "route"}),
    "Sole-Measured Variables": r5.rename(columns={
        "variable": "model-produced variable",
        "measurerName": "sole measurer (as named in the paper)",
        "k": "match key",
        "aliasFamily": "alias-resolved instrument family"}),
    "Sole-Measured by Instrument": sci[sci["R5_nSoleVars"] > 0][
        ["instr", "rank", "criticality", "R5_nSoleVars", "R5_soleVars"]]
        .rename(columns={"instr": "instrument",
                         "R5_nSoleVars": "sole-measured variables (strict GCMD join)",
                         "R5_soleVars": "variable list"}),
    "Study Regions": reg.sort_values("nPapers", ascending=False),
    "Country Mentions": geo.sort_values("nPapers", ascending=False),
    "Boundary Cohort": coh.sort_values("nAuthorOrcids", ascending=False),
    "Verified Quantities": kv,
    "Methods & Rules": methods,
}

TIER_FILL = {"A": PatternFill("solid", fgColor="C6E0B4"),
             "B": PatternFill("solid", fgColor="FFE699"),
             "C": PatternFill("solid", fgColor="F2F2F2")}
HDR = PatternFill("solid", fgColor="1F4E79")

with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    for name, dframe in sheets.items():
        dframe.to_excel(xw, sheet_name=name[:31], index=False)
        ws = xw.sheets[name[:31]]
        for c in range(1, dframe.shape[1] + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = HDR
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            width = max(11, min(48, int(dframe.iloc[:, c - 1]
                                        .astype(str).str.len().max() or 11) + 2))
            ws.column_dimensions[get_column_letter(c)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                if name == "Methods & Rules":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        if name == "Ranked Results":
            tcol = list(dframe.columns).index("tier") + 1
            for r in range(2, len(dframe) + 2):
                t = ws.cell(row=r, column=tcol).value
                if t in TIER_FILL:
                    for c in range(1, dframe.shape[1] + 1):
                        ws.cell(row=r, column=c).fill = TIER_FILL[t]
        if name == "Methods & Rules":
            ws.column_dimensions["A"].width = 26
            ws.column_dimensions["B"].width = 110

print(f"wrote {OUT}  ({len(sheets)} sheets, "
      f"{len(ranked)} ranked instruments)")
