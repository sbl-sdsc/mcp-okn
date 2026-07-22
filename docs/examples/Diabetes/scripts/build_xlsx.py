import pandas as pd, json, numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
D='data'; S=json.load(open('stats.json'))
rd=lambda f,**k: pd.read_csv(D+'/'+f,low_memory=False,**k)
out='Diabetes_results.xlsx'
sheets={}
g=rd('ranked_genes_tiered.csv')
sheets['Ranked Results']=g
sheets['GO enrichment']=rd('enrichment_GO.csv')
sheets['Reactome enrichment']=rd('enrichment_Reactome.csv')
sheets['Trait gene-set enrichment']=rd('enrichment_trait_broad.csv')
sheets['Non-coding RNA loci']=(lambda b: b[b.entrez.isna()][['symbol','rsid']].drop_duplicates().sort_values(['symbol','rsid']))(rd('biomarkerkg_t2d_risk_variants.csv'))
sheets['Risk variants (coding)']=(lambda b: b[b.entrez.notna()][['symbol','entrez','rsid','rel']].drop_duplicates())(rd('biomarkerkg_t2d_risk_variants.csv'))
ph=rd('oard_t2d_phenotypes.csv')
ph=ph[np.isfinite(pd.to_numeric(ph.log_odds_ratio,errors='coerce'))]
sheets['Phenotypes (oard-kg)']=ph.sort_values('log_odds_ratio',ascending=False).head(400)[
  ['partner_label','partner_type','t2d_position','concept_pair_count','log_odds_ratio','log_odds_ratio_ci_low','log_odds_ratio_ci_high','odds_ratio','dataset','total_sample_size']]
rl=rd('rdkg_t2d_relations.csv')
sheets['Drugs (rdkg treats)']=rl[rl.rel.isin(['treats','contraindicated_for'])][['dLabel','rel','objLabel','obj']].drop_duplicates()
sheets['Indications (prokn)']=rd('prokn_t2d_indications.csv').drop_duplicates()
sheets['Drug targets (prokn)']=rd('prokn_t2d_drug_targets.csv')
tc=rd('prokn_core_target_compounds.csv')
sheets['Target-anchored candidates']=tc[tc.interaction_predicate.astype(str).str.contains('RO_0002436')]
sheets['Repurposing shortlist']=rd('repurposing_candidates.csv').head(1000)
sheets['Islet expression (GXA)']=rd('gxa_t2d_expression.csv')
sheets['Islet chromatin (pankgraph)']=rd('pankgraph_t2d_ocr_top_contrasts.csv')
sheets['Exposure-gene (ICE tox)']=rd('biobricks_tox_t2d_genes.csv')
sheets['Adverse outcome pathways']=rd('aopwiki_t2d_aops.csv')
sheets['County prevalence + SDoH']=rd('county_analysis_matrix.csv',dtype={'county_FIPS':str})
sheets['SDoH correlations']=rd('county_sdoh_correlations.csv')
sheets['Multivariable model']=rd('county_multivariable_model.csv')
sheets['Diabetes complications']=rd('mondo_diabetes_complications.csv')
sheets['KG reconciliation']=rd('kg_reconciliation.csv')
methods=pd.DataFrame({'Item':[
 'Study','Anchor disease','Disease scope','Endpoint','Consensus universe','Consensus core','Tier A / B / C',
 'Evidence streams (disease-anchored)','Enrichment method','GO background','Reactome background',
 'Trait gene-set test','Epidemiology outcome','Epidemiology model','Level of inference','Abbreviations'],
 'Value':[
 'Multi-knowledge-graph integrative map of Type 2 Diabetes over the OKN federated SPARQL endpoint',
 'MONDO:0005148 (type 2 diabetes mellitus); crosswalks DOID:9352 · EFO:0001360 · OMIM:125853 · UMLS:C0011860 · MeSH:D003924 · SNOMED 44054006 · NCIT:C26747 · ICD-10-CM E11',
 f'{S["mondo_t2d_subtree"]}-term MONDO T2D subtree; {S["mondo_dm_subtree"]}-term diabetes-mellitus superclass used for context',
 'OKN federated SPARQL via mcp-okn',
 f'{S["genes_universe"]} genes with >=1 disease-anchored evidence stream',
 f'{S["genes_core"]} genes with >=2 disease-anchored evidence streams',
 f'{S["tier_a"]} / {S["tier_b"]} / {S["tier_c"]}',
 'digcfdekg PIGEAN gene-to-trait (statistical); spoke-okn ASSOCIATES_DaG (curated); biomarkerkg dbSNP risk variants (genetic); prokn ClinVar associated_with (curated)',
 'Hypergeometric over-representation with Benjamini-Hochberg FDR; k>=4, K>=3',
 f'{S["go_N"]} ProKN genes carrying >=1 GO annotation',
 f'{S["rx_N"]} ProKN genes carrying >=1 human Reactome pathway',
 f'Broad (digcfdekg PIGEAN trait sets) vs a digcfdekg-independent signature (n={S["trait_sig_n"]}); curated (rdkg neonatal-diabetes Mendelian set, K=5)',
 'County Health Rankings adult diagnosed-diabetes prevalence via spoke-okn PREVALENCEIN_SpL',
 f'OLS on standardized predictors, n={S["model_n"]} counties, R2={S["model_r2"]}',
 'Hypothesis generation. Association edges are observational, not causal; EHR co-occurrence is not causal; county-level results are ecological.',
 'AOP=adverse outcome pathway; BH=Benjamini-Hochberg; CGM=continuous glucose monitor; CL=Cell Ontology; DE=differential expression; eQTL=expression quantitative trait locus; FDR=false-discovery rate; FIPS=Federal Information Processing Standards county code; GO=Gene Ontology; GWAS=genome-wide association study; HP=Human Phenotype Ontology; ICE=Integrated Chemical Environment; KG=knowledge graph; lncRNA=long non-coding RNA; MoA=mechanism of action; OCR=open chromatin region; OLS=ordinary least squares; PFAS=per- and polyfluoroalkyl substances; PIP=posterior inclusion probability; SDoH=social determinants of health; T1D=type 1 diabetes; T2D=type 2 diabetes; UBERON=Uber-anatomy ontology; YPLL=years of potential life lost']})
sheets['Methods & Rules']=methods
with pd.ExcelWriter(out,engine='openpyxl') as w:
    for name,df in sheets.items():
        df.to_excel(w,sheet_name=name[:31],index=False)
    wb=w.book
    hdr=PatternFill('solid',fgColor='1F4E79'); hf=Font(name='Arial',size=10,bold=True,color='FFFFFF')
    tier={'A':PatternFill('solid',fgColor='FCE4D6'),'B':PatternFill('solid',fgColor='FFF2CC'),'C':PatternFill('solid',fgColor='F2F2F2')}
    for name,df in sheets.items():
        ws=wb[name[:31]]
        for c in ws[1]: c.fill=hdr; c.font=hf; c.alignment=Alignment(wrap_text=True,vertical='center')
        ws.freeze_panes='A2'
        if len(df): ws.auto_filter.ref=ws.dimensions
        for i,col in enumerate(df.columns,1):
            wmax=max(len(str(col)), *(len(str(v)) for v in df[col].head(400).astype(str))) if len(df) else len(str(col))
            ws.column_dimensions[get_column_letter(i)].width=min(max(10,wmax+2),58)
        for r in ws.iter_rows(min_row=2):
            for c in r: c.font=Font(name='Arial',size=10)
        if name=='Ranked Results':
            ti=list(df.columns).index('tier')+1
            for r in range(2,min(len(df)+2,3000)):
                t=ws.cell(row=r,column=ti).value
                if t in tier:
                    for c in range(1,len(df.columns)+1): ws.cell(row=r,column=c).fill=tier[t]
print('wrote',out,'sheets:',len(sheets))
