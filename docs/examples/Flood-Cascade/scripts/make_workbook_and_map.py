import sys, json, pandas as pd, numpy as np
sys.path.insert(0,"/sessions/affectionate-festive-ramanujan/mnt/.claude/skills/okn-report-style/scripts")
from okn_figstyle import folium_osm_map, save_map_html
T=pd.read_csv('data/county_typology.csv', dtype={'fips':str})
cen=pd.read_csv('data/county_centroids.csv', dtype={'fips':str}); T=T.merge(cen,on='fips',how='left')
S=json.load(open('data/headline_stats.json'))
T['sources']=T.apply(lambda r: [s for s,ok in [('ufokn',r.flood_cells>0),('fiokg',r.local_fac>0),
    ('hydrologykg',r.imported_fac>0 or r.ds_reaches>0),('spatialkg',True),('ruralkg',pd.notna(r.rucc)),
    ('sawgraph',r.ds_monitored_cells>0)] if ok], axis=1)
T['n_sources']=T.sources.apply(len)
T['state']=T.fips.str[:2]
ST={'01':'AL','04':'AZ','05':'AR','06':'CA','08':'CO','09':'CT','10':'DE','12':'FL','13':'GA','16':'ID','17':'IL','18':'IN','19':'IA','20':'KS','21':'KY','22':'LA','23':'ME','24':'MD','25':'MA','26':'MI','27':'MN','28':'MS','29':'MO','30':'MT','31':'NE','32':'NV','33':'NH','34':'NJ','35':'NM','36':'NY','37':'NC','38':'ND','39':'OH','40':'OK','41':'OR','42':'PA','44':'RI','45':'SC','46':'SD','47':'TN','48':'TX','49':'UT','50':'VT','51':'VA','53':'WA','54':'WV','55':'WI','56':'WY','11':'DC'}
T['st']=T.state.map(ST)
T.to_csv('data/county_typology.csv',index=False)

# ---------- Excel ----------
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
cols=['fips','label','st','typology','strongest_tier','local_fac','imported_fac','imported_fac_wt',
      'upstream_src_reaches','upstream_src_counties','ds_reaches','ds_monitored_cells','flood_cells',
      'flood_buildings','max_depth_m','rucc','pop','rural','retained_score','imported_score',
      'cascade_score','baseline_rank','cascade_rank','rank_shift','n_sources']
R=T[cols].copy()
sheets={
 'Ranked Results':R,
 'Industry sectors':pd.read_csv('data/industry_sectors.csv'),
 'Industry NAICS-6':pd.read_csv('data/industry_naics6.csv'),
 'EPA programs':pd.read_csv('data/epa_programs.csv'),
 'Flooded wells':pd.read_csv('data/flooded_wells.csv'),
 'Routing links (sample)':pd.read_csv('data/routing_links_tiered.csv.gz', dtype=str).head(50000),
}
meth=pd.DataFrame({'Item':[
 'Study','Unit of analysis','Flood footprint','Source inventory','Routing','Tier A','Tier B','Tier C',
 'Imported score','Retained score','Typology cut','Baseline','Direct pathway','Abbreviations'],
 'Value':[
 'Flood-Cascade: following flood-mobilised contamination downstream through the OKN federation',
 'US county (5-digit FIPS); intermediate units = S2 Level-13 cells and NHDPlus reaches (COMID)',
 'UF-OKN modelled flood-depth predictions for buildings; 47,512 buildings -> 2,738 S2 L13 cells',
 'EPA FRS facilities (fiokg) whose S2 L13 cell is a flood cell; PFAS/monitoring features (sawgraph); water wells (hydrologykg)',
 'hydrologykg downstreamFlowPathTC (NHDPlus transitive downstream closure) from reaches crossing a source flood cell',
 'downstream reach in the SAME HUC8 as the source (weight 1.0)',
 'same HUC4, different HUC8 (weight 0.5)',
 'same HUC2, different HUC4 (weight 0.25); Tier D = cross-region (weight 0.1; none observed)',
 'percentile rank of sum(tier-weighted upstream flood-exposed facilities) over source reaches upstream of the county',
 'percentile rank of flood-exposed facilities co-located inside the county',
 'high = percentile >= 0.60 on the respective axis',
 'ranking on co-located flood-exposed facility count alone (no routing)',
 'water wells (hydrologykg) whose S2 L13 cell is a flood cell',
 'FIPS = Federal Information Processing Standard county code; HUC = hydrologic unit code; NAICS = North American Industry Classification System; NHDPlus = National Hydrography Dataset Plus; COMID = NHDPlus common identifier; RUCC = Rural-Urban Continuum Code; PFAS = per- and polyfluoroalkyl substances; S2 = Google S2 geometry grid; HHI = Herfindahl-Hirschman Index; EPA FRS = Environmental Protection Agency Facility Registry Service']})
sheets['Methods & Rules']=meth
with pd.ExcelWriter('Flood-Cascade_results.xlsx', engine='openpyxl') as xl:
    for name,df in sheets.items(): df.to_excel(xl, sheet_name=name[:31], index=False)
wb=openpyxl.load_workbook('Flood-Cascade_results.xlsx')
FILL={'Compound':'FFD9C6','Imported':'FFF0CC','Retained':'D6E8F5','Low':'F2F2F2'}
for ws in wb.worksheets:
    ws.freeze_panes='A2'
    ws.auto_filter.ref=ws.dimensions
    for c in ws[1]:
        c.font=Font(name='Arial', bold=True, color='FFFFFF'); c.fill=PatternFill('solid', fgColor='2F5597')
        c.alignment=Alignment(wrap_text=True, vertical='center')
    for col in ws.columns:
        w=max((len(str(c.value)) for c in col[:200] if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(w+2,10),42)
    for row in ws.iter_rows(min_row=2):
        for c in row: c.font=Font(name='Arial', size=10)
ws=wb['Ranked Results']; ti=cols.index('typology')+1
for row in ws.iter_rows(min_row=2):
    t=row[ti-1].value
    if t in FILL:
        for c in row: c.fill=PatternFill('solid', fgColor=FILL[t])
wb.save('Flood-Cascade_results.xlsx')
print("xlsx ok")

# ---------- folium interactive map ----------
COL={'Compound':'#D55E00','Imported':'#E69F00','Retained':'#0072B2','Low':'#BBBBBB'}
rows=[]
for _,r in T.dropna(subset=['lat','lng']).iterrows():
    rows.append(dict(lat=float(r.lat), lon=float(r.lng), county=r.label, typology=r.typology,
        co_located_facilities=int(r.local_fac), imported_upstream_facilities=int(r.imported_fac),
        upstream_counties=int(r.upstream_src_counties), rucc=(None if pd.isna(r.rucc) else int(r.rucc)),
        population=(None if pd.isna(r['pop']) else int(r['pop'])),
        downstream_monitoring_cells=int(r.ds_monitored_cells), sources=", ".join(r.sources)))
m=folium_osm_map(rows, popup_keys=['county','typology','co_located_facilities','imported_upstream_facilities',
    'upstream_counties','rucc','population','downstream_monitoring_cells','sources'],
    tooltip_key='county')
import folium
for r in rows:
    folium.CircleMarker([r['lat'],r['lon']], radius=6 if r['typology']=='Compound' else 4,
        color=COL[r['typology']], fill=True, fill_opacity=.85, weight=1,
        popup=folium.Popup("<br>".join(f"<b>{k}</b>: {v}" for k,v in r.items() if k not in ('lat','lon')), max_width=320),
        tooltip=r['county']).add_to(m)
save_map_html(m,'data/typology_map.html')
print("map ok")
